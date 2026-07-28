from __future__ import annotations

from dataclasses import asdict, dataclass, field
from io import BytesIO
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class PackageComponent:
    component_type: str
    manufacturer: str = ""
    model: str = ""
    description: str = ""
    quantity: float = 1
    supplier: str = ""
    product_link: str = ""
    unit_price: float | None = None
    lead_time: str = ""
    stock_status: str = ""
    quote_number: str = ""
    notes: str = ""

    def to_row(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class FixturePackage:
    item_tag: str
    package_name: str
    quantity: float = 1
    source_description: str = ""
    components: list[PackageComponent] = field(default_factory=list)
    notes: str = ""

    def to_row(self) -> dict[str, Any]:
        return {
            "item_tag": self.item_tag,
            "package_name": self.package_name,
            "quantity": self.quantity,
            "component_count": len(self.components),
            "source_description": self.source_description,
            "notes": self.notes,
        }


S1_COMPONENTS = [
    PackageComponent("Sink", "JUST", "USXN1842A-J", "Single-compartment 18-gauge Type 304 stainless-steel sink, self-rimming, 26-1/2 x 18-1/2 x 10 in."),
    PackageComponent("Faucet", "Chicago Faucets", "350-GN8AE35ABCP", "1.5 GPM faucet"),
    PackageComponent("Bubbler", "Chicago Faucets", "748-665ABCP", "Bubbler"),
    PackageComponent("Angle stop", "Chicago Faucets", "1013-ABCP", "Angle stop with loose key; McGuire LFCK09LK permitted by schedule"),
    PackageComponent("Basket strainer", "McGuire", "152N", "Flat basket strainer"),
    PackageComponent("P-trap insulation kit", "McGuire", "PW2150GJ", "1-1/2 in. P-trap kit with ground joint and seamless pre-wrapped insulation"),
]


def _copy_component(component: PackageComponent) -> PackageComponent:
    return PackageComponent(**component.to_row())


def build_known_package(item_tag: str, description: str, quantity: float = 1) -> FixturePackage | None:
    """Recognize known multi-component fixture packages from schedule text."""
    text = re.sub(r"\s+", " ", description or "").strip()
    upper = text.upper()
    normalized_tag = (item_tag or "").strip().upper()
    if normalized_tag == "S-1" or ("USXN1842A-J" in upper and "350-GN8AE35ABCP" in upper):
        return FixturePackage(
            item_tag=normalized_tag or "S-1",
            package_name="S-1 Complete Sink Package",
            quantity=quantity,
            source_description=text,
            components=[_copy_component(c) for c in S1_COMPONENTS],
            notes="Buy as one vendor quote/PO when possible; components remain separate manufacturer SKUs.",
        )
    return None


def parse_component_lines(item_tag: str, package_name: str, text: str, quantity: float = 1) -> FixturePackage:
    """Build a package from one-component-per-line text.

    Accepted format: Component | Manufacturer | Model | Description | Qty
    Missing fields are allowed.
    """
    components: list[PackageComponent] = []
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split("|")]
        parts += [""] * (5 - len(parts))
        try:
            component_qty = float(parts[4]) if parts[4] else 1
        except ValueError:
            component_qty = 1
        components.append(PackageComponent(parts[0], parts[1], parts[2], parts[3], component_qty))
    return FixturePackage(item_tag.strip(), package_name.strip() or "Fixture Package", quantity, text, components)


def package_to_rfq_items(package: FixturePackage) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for component in package.components:
        rows.append({
            "include": True,
            "item_tag": package.item_tag,
            "manufacturer": component.manufacturer,
            "model": component.model,
            "description": f"{component.component_type}: {component.description}".strip(": "),
            "quantity": package.quantity * component.quantity,
            "product_link": component.product_link,
            "spec_link": "",
            "target_vendor": component.supplier,
            "notes": f"Part of {package.package_name}. {component.notes}".strip(),
        })
    return rows


def create_package_workbook(package: FixturePackage) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Package Summary"
    navy, blue, pale, green = "17324D", "2E75B6", "EAF3FB", "E2F0D9"
    thin = Side(style="thin", color="C8D5E2")

    ws.merge_cells("A1:N1")
    ws["A1"] = package.package_name.upper()
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A3"], ws["B3"] = "Item Tag", package.item_tag
    ws["D3"], ws["E3"] = "Package Qty", package.quantity
    ws["G3"], ws["H3"] = "Component Count", len(package.components)
    ws["A5"] = "Source / Schedule Description"
    ws["A5"].font = Font(bold=True)
    ws.merge_cells("B5:N6")
    ws["B5"] = package.source_description
    ws["B5"].alignment = Alignment(wrap_text=True, vertical="top")

    headers = ["Line", "Component", "Manufacturer", "Model / MPN", "Description", "Qty", "Supplier", "Unit Price", "Extended", "Stock Status", "Lead Time", "Quote #", "Product Link", "Notes"]
    start = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, component in enumerate(package.components, 1):
        row = start + idx
        qty = package.quantity * component.quantity
        values = [idx, component.component_type, component.manufacturer, component.model, component.description, qty, component.supplier, component.unit_price or "", f'=IFERROR(F{row}*H{row},0)', component.stock_status, component.lead_time, component.quote_number, component.product_link, component.notes]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if idx % 2:
                cell.fill = PatternFill("solid", fgColor=pale)
            if col in (8, 9):
                cell.number_format = '$#,##0.00;[Red]($#,##0.00);-'
        if component.product_link:
            ws.cell(row, 13).hyperlink = component.product_link
            ws.cell(row, 13).style = "Hyperlink"
    total_row = start + len(package.components) + 2
    ws.cell(total_row, 8, "Package Total").font = Font(bold=True)
    ws.cell(total_row, 9, f"=SUM(I{start+1}:I{start+len(package.components)})")
    ws.cell(total_row, 9).font = Font(bold=True)
    ws.cell(total_row, 9).fill = PatternFill("solid", fgColor=green)
    ws.cell(total_row, 9).number_format = '$#,##0.00;[Red]($#,##0.00);-'

    widths = [7, 20, 20, 22, 46, 8, 22, 13, 14, 16, 18, 16, 36, 40]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = f"A{start+1}"
    ws.auto_filter.ref = f"A{start}:N{start+len(package.components)}"
    ws.sheet_view.showGridLines = False

    vendor = wb.create_sheet("Vendor Quote Return")
    vendor.append(["Package", "Item Tag", "Vendor", "Quote #", "Quote Date", "Quote Expires", "Package Price", "Freight", "Tax", "Lead Time", "Earliest Ship Date", "Stock Status", "Notes"])
    vendor.append([package.package_name, package.item_tag, "", "", "", "", "", "", "", "", "", "", ""])
    for cell in vendor[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(wrap_text=True)
    for col in range(1, 14):
        vendor.column_dimensions[get_column_letter(col)].width = 20 if col not in (1, 13) else 36
    for col in (7, 8, 9):
        vendor.cell(2, col).number_format = '$#,##0.00;[Red]($#,##0.00);-'
    vendor.freeze_panes = "A2"
    vendor.sheet_view.showGridLines = False

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
