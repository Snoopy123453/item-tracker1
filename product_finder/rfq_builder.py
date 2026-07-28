from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import date
from io import BytesIO
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class RFQItem:
    include: bool = True
    item_tag: str = ""
    manufacturer: str = ""
    model: str = ""
    description: str = ""
    quantity: float = 1
    product_link: str = ""
    spec_link: str = ""
    target_vendor: str = ""
    notes: str = ""

    def to_row(self) -> dict[str, Any]:
        return asdict(self)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _find_header_row(ws, candidates: set[str], max_rows: int = 20) -> tuple[int, dict[str, int]] | None:
    for row in range(1, min(ws.max_row, max_rows) + 1):
        mapping: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            text = _clean(ws.cell(row, col).value).lower().replace("_", " ")
            if text:
                mapping[text] = col
        if any(key in mapping for key in candidates):
            return row, mapping
    return None


def extract_rfq_items(workbook_bytes: bytes) -> list[RFQItem]:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=False)
    preferred = [
        "Equipment Register", "Products", "Product Results", "Best Matches",
        "Approved Products", "Purchase List", "OmniSearch Results",
    ]
    sheets = [name for name in preferred if name in wb.sheetnames] + [n for n in wb.sheetnames if n not in preferred]
    output: list[RFQItem] = []
    seen: set[tuple[str, str, str]] = set()

    aliases = {
        "item_tag": ["item", "item tag", "tag"],
        "manufacturer": ["manufacturer", "brand"],
        "model": ["model", "model number", "mpn", "part number"],
        "description": ["description", "product", "title", "fixture", "extracted product name"],
        "quantity": ["quantity", "qty", "total quantity"],
        "product_link": ["product link", "retailer page", "link", "approved product link", "seller link"],
        "spec_link": ["spec link", "spec sheet", "document link", "official link"],
        "target_vendor": ["seller", "vendor", "retailer", "source name"],
    }

    for sheet_name in sheets:
        ws = wb[sheet_name]
        header = _find_header_row(ws, {a for values in aliases.values() for a in values})
        if not header:
            continue
        header_row, mapping = header

        def col_for(field: str) -> int | None:
            for alias in aliases[field]:
                if alias in mapping:
                    return mapping[alias]
            return None

        cols = {field: col_for(field) for field in aliases}
        if not any(cols[k] for k in ("model", "description", "product_link")):
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            values = {field: _clean(ws.cell(r, col).value) if col else "" for field, col in cols.items()}
            if not any(values[k] for k in ("model", "description", "product_link")):
                continue
            qty_raw = values.get("quantity") or "1"
            try:
                qty = float(str(qty_raw).replace(",", ""))
            except ValueError:
                qty = 1
            link = values.get("product_link", "")
            if not link:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(r, col)
                    if cell.hyperlink and cell.hyperlink.target:
                        link = cell.hyperlink.target
                        break
            key = (values.get("manufacturer", "").lower(), values.get("model", "").lower(), link.lower())
            if key in seen:
                continue
            seen.add(key)
            output.append(RFQItem(
                item_tag=values.get("item_tag", ""), manufacturer=values.get("manufacturer", ""),
                model=values.get("model", ""), description=values.get("description", ""),
                quantity=max(qty, 1), product_link=link, spec_link=values.get("spec_link", ""),
                target_vendor=values.get("target_vendor", ""),
            ))
        if output and sheet_name in preferred:
            break
    return output


def build_rfq_email(project_name: str, ship_to: str, needed_by: str, contact_name: str,
                    contact_email: str, items: list[dict[str, Any]], substitutions: str,
                    tax_exempt: bool, notes: str = "") -> str:
    lines = [
        f"Subject: Request for Quote - {project_name or 'Product Procurement'}",
        "",
        "Hello,",
        "",
        "Please provide a formal quote and current lead time for the items below.",
        f"Ship-to location: {ship_to or 'To be confirmed'}",
        f"Requested delivery date: {needed_by or 'Please advise earliest availability'}",
        f"Substitutions: {substitutions}",
        f"Tax-exempt purchase: {'Yes' if tax_exempt else 'No'}",
        "",
        "Items:",
    ]
    for idx, item in enumerate(items, 1):
        identity = " ".join(x for x in [_clean(item.get("manufacturer")), _clean(item.get("model"))] if x)
        desc = _clean(item.get("description"))
        lines.append(f"{idx}. {identity or desc} | Qty: {item.get('quantity', 1):g}" + (f" | {desc}" if identity and desc else ""))
        if _clean(item.get("product_link")):
            lines.append(f"   Reference: {_clean(item.get('product_link'))}")
        if _clean(item.get("spec_link")):
            lines.append(f"   Spec: {_clean(item.get('spec_link'))}")
    lines.extend([
        "",
        "Please include unit price, freight, taxes, stock status, manufacturer lead time, estimated ship date, quote expiration, and any proposed substitutions.",
    ])
    if notes:
        lines.extend(["", f"Additional notes: {notes}"])
    lines.extend(["", "Thank you,", contact_name or "Purchasing", contact_email])
    return "\n".join(lines)


def create_rfq_workbook(project: dict[str, Any], items: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ"
    dark = "17324D"; blue = "2E75B6"; light = "EAF3FB"; green = "E2F0D9"; gray = "666666"
    thin = Side(style="thin", color="C8D5E2")

    ws.merge_cells("A1:N1")
    ws["A1"] = "REQUEST FOR QUOTE"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    fields = [
        ("Project", project.get("project_name", "")), ("Project Number", project.get("project_number", "")),
        ("Ship To", project.get("ship_to", "")), ("Needed By", project.get("needed_by", "")),
        ("Contact", project.get("contact_name", "")), ("Email", project.get("contact_email", "")),
        ("Phone", project.get("contact_phone", "")), ("Substitutions", project.get("substitutions", "")),
        ("Tax Exempt", "Yes" if project.get("tax_exempt") else "No"), ("RFQ Date", date.today().isoformat()),
    ]
    for i, (label, value) in enumerate(fields):
        row = 3 + i // 2
        col = 1 if i % 2 == 0 else 8
        ws.cell(row, col, label).font = Font(bold=True, color=gray)
        ws.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + 5)
        ws.cell(row, col + 1, value)

    start = 9
    headers = ["Line", "Item Tag", "Manufacturer", "Model / MPN", "Description", "Qty", "Unit Price", "Freight", "Tax", "Total", "Stock Status", "Lead Time", "Quote Expires", "Product / Spec Link"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start, c, h); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=blue); cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, item in enumerate(items, 1):
        r = start + idx
        values = [idx, item.get("item_tag", ""), item.get("manufacturer", ""), item.get("model", ""), item.get("description", ""), item.get("quantity", 1), "", "", "", f"=IFERROR(F{r}*G{r}+H{r}+I{r},0)", "", "", "", item.get("product_link") or item.get("spec_link") or ""]
        for c, value in enumerate(values, 1):
            cell = ws.cell(r, c, value); cell.alignment = Alignment(vertical="top", wrap_text=True)
            if idx % 2 == 1: cell.fill = PatternFill("solid", fgColor=light)
            if c in (7,8,9,10): cell.number_format = '$#,##0.00;[Red]($#,##0.00);-'
        link = values[-1]
        if link:
            ws.cell(r, 14).hyperlink = link; ws.cell(r, 14).style = "Hyperlink"
        ws.row_dimensions[r].height = 42

    total_row = start + len(items) + 2
    ws.cell(total_row, 9, "Quoted Total").font = Font(bold=True)
    ws.cell(total_row, 10, f"=SUM(J{start+1}:J{start+len(items)})").font = Font(bold=True)
    ws.cell(total_row, 10).fill = PatternFill("solid", fgColor=green)
    ws.cell(total_row, 10).number_format = '$#,##0.00;[Red]($#,##0.00);-'

    for row in ws.iter_rows(min_row=start, max_row=total_row, min_col=1, max_col=14):
        for cell in row: cell.border = Border(bottom=thin)
    widths = [7,12,20,20,42,8,13,13,12,14,16,16,15,36]
    for i, width in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = f"A{start+1}"; ws.auto_filter.ref = f"A{start}:N{start+len(items)}"; ws.sheet_view.showGridLines = False

    email_ws = wb.create_sheet("Email Draft")
    email_ws["A1"] = "EMAIL-READY RFQ"; email_ws["A1"].font = Font(size=18, bold=True, color="FFFFFF"); email_ws["A1"].fill = PatternFill("solid", fgColor=dark)
    email_ws.column_dimensions["A"].width = 120
    email_ws["A3"] = project.get("email_draft", ""); email_ws["A3"].alignment = Alignment(wrap_text=True, vertical="top"); email_ws.row_dimensions[3].height = 420
    email_ws.sheet_view.showGridLines = False

    out = BytesIO(); wb.save(out); return out.getvalue()
