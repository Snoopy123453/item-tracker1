from __future__ import annotations

from datetime import date
from io import BytesIO
import re
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .utils import safe_filename

NAVY = "17324D"
BLUE = "2E75B6"
LIGHT_BLUE = "DCEAF7"
PALE_BLUE = "EFF6FC"
GREEN = "008000"
GRAY = "666666"
WHITE = "FFFFFF"
ORANGE = "FCE4D6"
TEAL = "DDEBF7"
LIGHT_RED = "F4CCCC"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color=WHITE, bold=True)
TITLE_FONT = Font(size=20, bold=True, color=NAVY)
THIN_BLUE = Side(style="thin", color=BLUE)
INVALID_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _safe(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    cleaned = INVALID_XML_CHARS.sub("", value)
    if cleaned.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + cleaned
    return cleaned


def _is_url(value: str) -> bool:
    try:
        parsed = urlparse(value.strip())
    except ValueError:
        return False
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def extract_purchase_candidates(workbook_bytes: bytes) -> list[dict[str, Any]]:
    """Extract purchasable listing rows and hyperlinks from a Product Hunter workbook."""
    wb = load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=False)
    preferred = ["Product Results", "Products", "Retailers"]
    sheet_names = preferred + [name for name in wb.sheetnames if name not in preferred]
    candidates: list[dict[str, Any]] = []
    seen: set[str] = set()

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        headers = {
            str(cell.value).strip().lower(): cell.column
            for cell in ws[1]
            if cell.value is not None and str(cell.value).strip()
        }
        link_col = next((headers.get(key) for key in ("product_link", "retailer page", "link", "url", "website") if headers.get(key)), None)
        title_col = next((headers.get(key) for key in ("title", "product", "product name", "item") if headers.get(key)), None)
        seller_col = next((headers.get(key) for key in ("seller", "retailer", "vendor", "store") if headers.get(key)), None)
        price_col = next((headers.get(key) for key in ("extracted_price", "price", "unit price") if headers.get(key)), None)
        image_col = next((headers.get(key) for key in ("thumbnail", "image", "image url") if headers.get(key)), None)
        query_col = next((headers.get(key) for key in ("query", "search term", "model") if headers.get(key)), None)

        for row_idx in range(2, ws.max_row + 1):
            link = ""
            if link_col:
                cell = ws.cell(row_idx, link_col)
                link = str(cell.hyperlink.target if cell.hyperlink else (cell.value or "")).strip()
            if not _is_url(link):
                for cell in ws[row_idx]:
                    possible = str(cell.hyperlink.target if cell.hyperlink else (cell.value or "")).strip()
                    if _is_url(possible):
                        link = possible
                        break
            if not _is_url(link) or link.casefold() in seen:
                continue
            seen.add(link.casefold())

            price_value = ws.cell(row_idx, price_col).value if price_col else ""
            try:
                unit_price = float(price_value) if price_value not in (None, "") else 0.0
            except (TypeError, ValueError):
                match = re.search(r"\d[\d,]*(?:\.\d{1,2})?", str(price_value or ""))
                unit_price = float(match.group(0).replace(",", "")) if match else 0.0

            candidates.append({
                "select": False,
                "product": str(ws.cell(row_idx, title_col).value or "") if title_col else "",
                "model_or_search": str(ws.cell(row_idx, query_col).value or "") if query_col else "",
                "retailer": str(ws.cell(row_idx, seller_col).value or "") if seller_col else "",
                "unit_price": unit_price,
                "quantity": 1,
                "product_link": link,
                "image_url": str(ws.cell(row_idx, image_col).value or "") if image_col else "",
                "source_sheet": sheet_name,
                "source_row": row_idx,
            })
    return candidates


def _style_data_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_BLUE)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = PatternFill("solid", fgColor=PALE_BLUE) if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, str) and _is_url(cell.value):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
    for col_idx, cells in enumerate(ws.columns, start=1):
        max_len = min(max([len(str(c.value or "")) for c in cells] + [10]), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)
    ws.row_dimensions[1].height = 34


def suggest_purchase_tracker_filename(rows: list[dict[str, Any]], tracker_name: str = "") -> str:
    if tracker_name.strip():
        stem = tracker_name.strip()
    else:
        products = []
        seen = set()
        for row in rows:
            name = str(row.get("model_or_search") or row.get("product") or "").strip()
            if name and name.casefold() not in seen:
                seen.add(name.casefold())
                products.append(name)
        if len(products) == 1:
            stem = f"{products[0]} Purchase Tracker"
        elif len(products) == 2:
            stem = f"{products[0]} and {products[1]} Purchase Tracker"
        else:
            stem = f"Purchase Tracker {len(rows)} Items"
    return safe_filename(f"{stem[:120]}.xlsx")


def create_purchase_tracker_bytes(rows: list[dict[str, Any]], tracker_name: str = "", project_name: str = "", buyer: str = "", notes: str = "") -> tuple[str, bytes]:
    wb = Workbook()
    summary = wb.active
    summary.title = "Dashboard"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:F1")
    summary["A1"] = (tracker_name.strip() or "Purchase Tracker").upper()
    summary["A1"].font = TITLE_FONT
    summary["A1"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 38

    info = [("Project", project_name or "Not specified"), ("Buyer", buyer or "Not specified"), ("Created", date.today().isoformat())]
    for idx, (label, value) in enumerate(info, start=3):
        summary[f"A{idx}"] = label
        summary[f"A{idx}"].font = HEADER_FONT
        summary[f"A{idx}"].fill = HEADER_FILL
        summary[f"B{idx}"] = _safe(value)
        summary.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=3)
        summary[f"B{idx}"].fill = PatternFill("solid", fgColor=TEAL)

    summary["A7"] = "Items"
    summary["A7"].font = HEADER_FONT
    summary["A7"].fill = HEADER_FILL
    summary["B7"] = len(rows)
    summary["C7"] = "Estimated total"
    summary["C7"].font = HEADER_FONT
    summary["C7"].fill = HEADER_FILL
    summary["D7"] = "='Purchase List'!N2"
    summary["D7"].number_format = '$#,##0.00;[Red]($#,##0.00);-'
    summary["E7"] = "Purchased total"
    summary["E7"].font = HEADER_FONT
    summary["E7"].fill = HEADER_FILL
    summary["F7"] = "='Purchase List'!N3"
    summary["F7"].number_format = '$#,##0.00;[Red]($#,##0.00);-'

    summary.merge_cells("A10:F10")
    summary["A10"] = "Tracker notes"
    summary["A10"].font = HEADER_FONT
    summary["A10"].fill = HEADER_FILL
    summary.merge_cells("A11:F14")
    summary["A11"] = _safe(notes or "Use the Purchase List sheet to update status, order number, dates, quantities, taxes, shipping, and received quantities.")
    summary["A11"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["A11"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    for col in range(1, 7):
        summary.column_dimensions[get_column_letter(col)].width = 20

    ws = wb.create_sheet("Purchase List")
    headers = [
        "item_id", "product", "model_or_search", "retailer", "product_link", "image_url",
        "quantity", "unit_price", "shipping", "tax", "estimated_total", "status",
        "order_number", "ordered_date", "expected_date", "received_date", "received_quantity",
        "payment_method", "purchaser", "department / cost code", "notes", "source_sheet", "source_row",
    ]
    ws.append(headers)
    for idx, row in enumerate(rows, start=1):
        ws.append([
            idx,
            _safe(row.get("product", "")),
            _safe(row.get("model_or_search", "")),
            _safe(row.get("retailer", "")),
            _safe(row.get("product_link", "")),
            _safe(row.get("image_url", "")),
            max(1, int(row.get("quantity") or 1)),
            float(row.get("unit_price") or 0),
            float(row.get("shipping") or 0),
            float(row.get("tax") or 0),
            f"=G{idx+1}*H{idx+1}+I{idx+1}+J{idx+1}",
            _safe(row.get("status") or "Planned"),
            _safe(row.get("order_number", "")),
            _safe(row.get("ordered_date", "")),
            _safe(row.get("expected_date", "")),
            _safe(row.get("received_date", "")),
            int(row.get("received_quantity") or 0),
            _safe(row.get("payment_method", "")),
            _safe(row.get("purchaser") or buyer),
            _safe(row.get("cost_code", "")),
            _safe(row.get("notes", "")),
            _safe(row.get("source_sheet", "")),
            row.get("source_row", ""),
        ])
    _style_data_sheet(ws)
    for row in range(2, ws.max_row + 1):
        for col in (8, 9, 10, 11):
            ws.cell(row, col).number_format = '$#,##0.00;[Red]($#,##0.00);-'
        status_cell = ws.cell(row, 12)
        if str(status_cell.value).lower() == "planned":
            status_cell.fill = PatternFill("solid", fgColor=ORANGE)

    # A compact totals block to the right of the list.
    ws["M1"] = "Purchase Summary"
    ws["M1"].fill = HEADER_FILL
    ws["M1"].font = HEADER_FONT
    ws["N1"] = "Value"
    ws["N1"].fill = HEADER_FILL
    ws["N1"].font = HEADER_FONT
    last = max(2, ws.max_row)
    ws["M2"] = "Estimated total"
    ws["N2"] = f"=SUM(K2:K{last})"
    ws["M3"] = "Purchased total"
    ws["N3"] = f'=SUMIF(L2:L{last},"Purchased",K2:K{last})+SUMIF(L2:L{last},"Ordered",K2:K{last})+SUMIF(L2:L{last},"Received",K2:K{last})'
    ws["M4"] = "Received total"
    ws["N4"] = f'=SUMIF(L2:L{last},"Received",K2:K{last})'
    ws["M5"] = "Open quantity"
    ws["N5"] = f"=SUM(G2:G{last})-SUM(Q2:Q{last})"
    for cell in ("N2", "N3", "N4"):
        ws[cell].number_format = '$#,##0.00;[Red]($#,##0.00);-'
    for row in range(2, 6):
        ws[f"M{row}"].fill = PatternFill("solid", fgColor=TEAL)
        ws[f"M{row}"].font = Font(bold=True, color=NAVY)
        ws[f"N{row}"].fill = PatternFill("solid", fgColor=PALE_BLUE)

    instructions = wb.create_sheet("Instructions")
    instructions.sheet_view.showGridLines = False
    instructions["A1"] = "PURCHASE TRACKER WORKFLOW"
    instructions["A1"].font = TITLE_FONT
    instructions.column_dimensions["A"].width = 110
    steps = [
        "1. Update quantity, unit price, shipping, and tax before ordering.",
        "2. Change status from Planned to Approved, Ordered, Purchased, Received, Backordered, or Cancelled.",
        "3. Add the order number, expected date, purchaser, payment method, and department/cost code.",
        "4. Record received quantity and received date as shipments arrive.",
        "5. Use the Dashboard and Purchase Summary formulas to monitor committed and received spending.",
        "6. Verify all retailer links, model numbers, prices, stock, shipping, and return terms before purchasing.",
    ]
    for idx, step in enumerate(steps, start=3):
        instructions[f"A{idx}"] = step
        instructions[f"A{idx}"].alignment = Alignment(wrap_text=True)
        instructions[f"A{idx}"].fill = PatternFill("solid", fgColor=PALE_BLUE if idx % 2 else WHITE)
        instructions.row_dimensions[idx].height = 30

    wb.properties.title = tracker_name.strip() or "Purchase Tracker"
    wb.properties.subject = "Product purchasing and receipt tracking"
    wb.properties.creator = "Product Hunter Pro"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return suggest_purchase_tracker_filename(rows, tracker_name), buffer.getvalue()
