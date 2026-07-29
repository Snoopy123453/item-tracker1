from __future__ import annotations

import json
import re
from dataclasses import dataclass, asdict
from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable
from urllib.parse import urlparse

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import clean_text, safe_filename

VENDOR_ALIASES = {
    "the home depot": "Home Depot", "home depot pro": "Home Depot", "homedepot.com": "Home Depot",
    "amazon.com": "Amazon", "amazon marketplace": "Amazon", "lowes.com": "Lowe's", "lowe s": "Lowe's",
    "grainger industrial supply": "Grainger", "w.w. grainger": "Grainger", "ferguson enterprises": "Ferguson",
    "supplyhouse.com": "SupplyHouse", "zoro.com": "Zoro", "walmart.com": "Walmart",
}

OFFER_BASE_COLUMNS = [
    "title", "manufacturer", "model", "seller", "product_link", "quantity",
    "unit_price", "shipping", "tax_rate", "discount", "accessory_cost",
    "match_score", "exact_model_match", "status", "approved",
    "lead_time_score", "authorized_distributor", "vendor_rating", "notes",
]


def normalize_offer_dataframe(frame: pd.DataFrame | None) -> pd.DataFrame:
    """Return a Streamlit data-editor-safe offer table with stable dtypes.

    CSV and Excel imports frequently infer numeric and boolean columns as object
    strings. Streamlit validates ``column_config`` against pandas dtypes and
    raises ``StreamlitAPIException`` when they disagree. This function supplies
    missing columns and coerces every configured field to the expected dtype.
    """
    products = pd.DataFrame() if frame is None else frame.copy()
    defaults: dict[str, Any] = {
        "quantity": 1.0, "unit_price": 0.0, "shipping": 0.0, "tax_rate": 0.0,
        "discount": 0.0, "accessory_cost": 0.0, "match_score": 0.0,
        "lead_time_score": 0.0, "vendor_rating": 0.0,
        "exact_model_match": False, "approved": False,
        "authorized_distributor": False, "status": "Needs review",
    }
    text_cols = ["title", "manufacturer", "model", "seller", "product_link", "status", "notes"]
    numeric_cols = ["quantity", "unit_price", "shipping", "tax_rate", "discount", "accessory_cost", "match_score", "lead_time_score", "vendor_rating"]
    bool_cols = ["exact_model_match", "approved", "authorized_distributor"]

    for col in OFFER_BASE_COLUMNS:
        if col not in products.columns:
            products[col] = defaults.get(col, "")
    for col in text_cols:
        products[col] = products[col].fillna("").astype(str)
    for col in numeric_cols:
        products[col] = pd.to_numeric(products[col], errors="coerce").fillna(defaults[col]).astype(float)

    truthy = {"1", "true", "yes", "y", "checked", "on"}
    for col in bool_cols:
        products[col] = products[col].map(
            lambda value: value if isinstance(value, bool) else str(value).strip().casefold() in truthy
        ).astype(bool)
    products["status"] = products["status"].replace("", "Needs review")
    return products[OFFER_BASE_COLUMNS]


DOC_TYPES = {
    "Specification Sheet": ("spec sheet", "specification", "technical data", "cut sheet"),
    "Submittal": ("submittal",), "Installation Manual": ("installation", "install guide", "instructions"),
    "O&M Manual": ("operation and maintenance", "o&m", "maintenance manual"),
    "Warranty": ("warranty",), "Parts List": ("parts", "exploded view"),
    "CAD/BIM": ("cad", "bim", "revit", "dwg"), "Safety": ("sds", "safety data"),
}

@dataclass
class Requirement:
    attribute: str
    required_value: str
    importance: str = "Required"  # Required, Preferred, Optional, Ignore
    weight: float = 1.0

@dataclass
class Comparison:
    attribute: str
    required_value: str
    found_value: str
    importance: str
    status: str
    notes: str = ""


def normalize_vendor(name: str, url: str = "") -> str:
    raw = clean_text(name).casefold()
    domain = urlparse(url).netloc.casefold().removeprefix("www.") if url else ""
    for alias, canonical in VENDOR_ALIASES.items():
        if alias in raw or alias in domain:
            return canonical
    if raw:
        return clean_text(name)
    return domain.split(".")[0].title() if domain else "Unknown Vendor"


def classify_document(title: str, link: str = "") -> str:
    text = f"{title} {link}".casefold()
    for label, terms in DOC_TYPES.items():
        if any(term in text for term in terms):
            return label
    return "Product Page" if not link.lower().endswith(".pdf") else "Technical Document"


def validate_document(link: str, requested_model: str = "", timeout: int = 8) -> dict[str, Any]:
    result = {"link": link, "opens": False, "is_pdf": False, "status_code": None, "model_in_url": False, "notes": ""}
    if not link.startswith(("http://", "https://")):
        result["notes"] = "Invalid URL"
        return result
    model_compact = re.sub(r"\W+", "", requested_model).casefold()
    result["model_in_url"] = bool(model_compact and model_compact in re.sub(r"\W+", "", link).casefold())
    try:
        response = requests.head(link, allow_redirects=True, timeout=(4, timeout), headers={"User-Agent":"ProductHunter/8.0"})
        if response.status_code in {403, 405}:
            response = requests.get(link, stream=True, timeout=(4, timeout), headers={"User-Agent":"ProductHunter/8.0"})
        result["status_code"] = response.status_code
        result["opens"] = response.ok
        content_type = response.headers.get("content-type", "").casefold()
        result["is_pdf"] = "pdf" in content_type or link.lower().split("?")[0].endswith(".pdf")
        if not response.ok:
            result["notes"] = f"HTTP {response.status_code}"
    except requests.RequestException as exc:
        result["notes"] = f"Request failed: {type(exc).__name__}"
    return result


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).casefold()).strip()


def compare_requirements(requirements: Iterable[dict[str, Any] | Requirement], listing: dict[str, Any]) -> tuple[list[Comparison], bool, float]:
    comparisons: list[Comparison] = []
    earned = 0.0
    possible = 0.0
    hard_reject = False
    searchable = _norm(" ".join(str(v) for v in listing.values() if v is not None))
    for raw in requirements:
        req = raw if isinstance(raw, Requirement) else Requirement(**{k: raw.get(k) for k in ("attribute","required_value","importance","weight") if k in raw})
        if req.importance == "Ignore" or not clean_text(req.required_value):
            continue
        found = clean_text(listing.get(req.attribute, ""))
        required_norm, found_norm = _norm(req.required_value), _norm(found)
        if not found_norm and required_norm in searchable:
            found = req.required_value
            found_norm = required_norm
        match = bool(required_norm and (required_norm == found_norm or required_norm in found_norm or found_norm in required_norm))
        if match:
            status, notes, points = "Match", "Confirmed", req.weight
        elif not found_norm:
            status, notes, points = "Not stated", "Manual verification required", 0.35 * req.weight
        else:
            status, notes, points = "Conflict", f"Found: {found}", 0.0
            if req.importance == "Required":
                hard_reject = True
        multiplier = {"Required": 3.0, "Preferred": 2.0, "Optional": 1.0}.get(req.importance, 1.0)
        earned += points * multiplier
        possible += req.weight * multiplier
        comparisons.append(Comparison(req.attribute, req.required_value, found, req.importance, status, notes))
    score = round(100 * earned / possible, 1) if possible else 0.0
    return comparisons, hard_reject, score


def package_completeness(required_components: Iterable[str], listing_text: str) -> dict[str, Any]:
    text = _norm(listing_text)
    required = [clean_text(x) for x in required_components if clean_text(x)]
    found = [x for x in required if _norm(x) in text]
    missing = [x for x in required if x not in found]
    percent = round(100 * len(found) / len(required), 1) if required else 100.0
    return {"required": required, "found": found, "missing": missing, "percent": percent, "label": f"{len(found)} of {len(required)} components"}


def landed_cost(unit_price: float, quantity: float = 1, shipping: float = 0, tax_rate: float = 0, discount: float = 0, accessory_cost: float = 0) -> dict[str, float]:
    subtotal = max(0.0, float(unit_price)) * max(0.0, float(quantity)) + max(0.0, float(accessory_cost))
    taxable = max(0.0, subtotal - max(0.0, float(discount)))
    tax = taxable * max(0.0, float(tax_rate))
    total = taxable + max(0.0, float(shipping)) + tax
    return {"subtotal": round(subtotal, 2), "tax": round(tax, 2), "delivered_total": round(total, 2)}


def duplicate_key(row: dict[str, Any]) -> str:
    upc = _norm(row.get("upc") or row.get("barcode"))
    model = _norm(row.get("model") or row.get("mpn"))
    manufacturer = _norm(row.get("manufacturer") or row.get("brand"))
    if upc: return f"upc:{upc}"
    if model: return f"model:{manufacturer}:{model}"
    title = _norm(row.get("title") or row.get("product") or row.get("description"))
    return f"title:{title[:120]}"


def group_duplicate_offers(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = duplicate_key(row)
        if key not in groups:
            groups[key] = {"duplicate_key": key, "product": row.get("title") or row.get("product") or row.get("description", ""), "manufacturer": row.get("manufacturer") or row.get("brand", ""), "model": row.get("model") or row.get("mpn", ""), "offers": [], "offer_count": 0}
        groups[key]["offers"].append(row)
        groups[key]["offer_count"] += 1
    return list(groups.values())


def data_health_checks(products: list[dict[str, Any]], documents: list[dict[str, Any]] | None = None) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    seen: set[str] = set()
    for idx, row in enumerate(products, 1):
        label = clean_text(row.get("title") or row.get("product") or row.get("description") or f"Row {idx}")
        key = duplicate_key(row)
        if key in seen: issues.append({"severity":"Warning","item":label,"issue":"Possible duplicate product"})
        seen.add(key)
        if not clean_text(row.get("model") or row.get("mpn")): issues.append({"severity":"Review","item":label,"issue":"Model number missing"})
        if not clean_text(row.get("seller") or row.get("vendor")): issues.append({"severity":"Review","item":label,"issue":"Vendor missing"})
        if row.get("quantity") in {0, "0"}: issues.append({"severity":"Error","item":label,"issue":"Quantity is zero"})
        if not clean_text(row.get("product_link") or row.get("link")): issues.append({"severity":"Review","item":label,"issue":"Product link missing"})
        if clean_text(row.get("status")).casefold() in {"alternate", "substitution"} and not row.get("approved"):
            issues.append({"severity":"Error","item":label,"issue":"Unapproved alternate"})
    for doc in documents or []:
        if not clean_text(doc.get("link")): issues.append({"severity":"Review","item":clean_text(doc.get("title")),"issue":"Document link missing"})
    return issues


def build_review_queue(products: list[dict[str, Any]], minimum_score: float = 85) -> list[dict[str, Any]]:
    queue = []
    for row in products:
        reasons=[]
        score=float(row.get("match_score") or 0)
        if score < minimum_score: reasons.append(f"Match score below {minimum_score:g}%")
        if not row.get("exact_model_match") and clean_text(row.get("model")): reasons.append("Exact model not confirmed")
        if clean_text(row.get("differences") or row.get("contradictions")): reasons.append("Conflicting attributes")
        if not clean_text(row.get("price") or row.get("unit_price")): reasons.append("Price unavailable")
        if reasons:
            out=dict(row); out["review_reasons"]="; ".join(reasons); queue.append(out)
    return queue


def vendor_score(row: dict[str, Any]) -> float:
    match=float(row.get("match_score") or 0)
    price_score=float(row.get("price_score") or 50)
    lead=float(row.get("lead_time_score") or 50)
    authorized=100 if row.get("authorized_distributor") else 50
    performance=float(row.get("vendor_rating") or 50)
    return round(.40*match + .20*price_score + .15*lead + .15*authorized + .10*performance, 1)


def create_procurement_control_workbook(project_name: str, products: list[dict[str, Any]], requirements: list[dict[str, Any]], documents: list[dict[str, Any]], audit_log: list[dict[str, Any]]) -> tuple[str, bytes]:
    wb=Workbook(); ws=wb.active; ws.title="Control Dashboard"
    navy="17324D"; blue="2E75B6"; white="FFFFFF"; pale="EAF3FB"; red="F4CCCC"; orange="FCE4D6"; green="D9EAD3"
    ws.sheet_view.showGridLines=False; ws.merge_cells("A1:H1"); ws["A1"]=(project_name or "PROCUREMENT CONTROL").upper(); ws["A1"].font=Font(size=20,bold=True,color=white); ws["A1"].fill=PatternFill("solid",fgColor=navy)
    review=build_review_queue(products); issues=data_health_checks(products,documents)
    metrics=[("Products",len(products)),("Needs Review",len(review)),("Data Issues",len(issues)),("Documents",len(documents))]
    for i,(label,value) in enumerate(metrics):
        col=1+i*2; ws.merge_cells(start_row=3,start_column=col,end_row=3,end_column=col+1); ws.cell(3,col,label).fill=PatternFill("solid",fgColor=blue); ws.cell(3,col).font=Font(bold=True,color=white)
        ws.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+1); ws.cell(4,col,value).fill=PatternFill("solid",fgColor=pale); ws.cell(4,col).font=Font(size=16,bold=True,color=navy)
    for c in range(1,9): ws.column_dimensions[get_column_letter(c)].width=18

    def add_sheet(name:str, rows:list[dict[str,Any]]):
        sh=wb.create_sheet(name); sh.sheet_view.showGridLines=False
        headers=list(dict.fromkeys(k for row in rows for k in row.keys())) if rows else ["status"]
        sh.append(headers)
        if not rows: sh.append(["No records"])
        for row in rows: sh.append([row.get(h,"") for h in headers])
        sh.freeze_panes="A2"; sh.auto_filter.ref=sh.dimensions
        for cell in sh[1]: cell.fill=PatternFill("solid",fgColor=navy); cell.font=Font(bold=True,color=white); cell.alignment=Alignment(wrap_text=True)
        for ci,h in enumerate(headers,1): sh.column_dimensions[get_column_letter(ci)].width=44 if h in {"description","product_link","link","differences","review_reasons","issue","notes"} else 18
        for row in sh.iter_rows(min_row=2):
            for cell in row:
                cell.alignment=Alignment(vertical="top",wrap_text=True)
                if isinstance(cell.value,str) and cell.value.startswith(("http://","https://")): cell.hyperlink=cell.value; cell.style="Hyperlink"
        return sh

    add_sheet("Products",products); add_sheet("Requirements",requirements); add_sheet("Review Queue",review); health=add_sheet("Data Health",issues); add_sheet("Documents",documents); add_sheet("Audit Log",audit_log)
    for row in health.iter_rows(min_row=2):
        severity=str(row[0].value)
        fill=red if severity=="Error" else orange if severity in {"Warning","Review"} else green
        for cell in row: cell.fill=PatternFill("solid",fgColor=fill)

    po_rows=[]
    for row in products:
        if clean_text(row.get("status")).casefold() in {"approved","selected","ordered"} or row.get("approved"):
            unit=float(row.get("unit_price") or row.get("extracted_price") or 0); qty=float(row.get("quantity") or 1); calc=landed_cost(unit,qty,float(row.get("shipping") or 0),float(row.get("tax_rate") or 0),float(row.get("discount") or 0),float(row.get("accessory_cost") or 0))
            po_rows.append({"vendor":normalize_vendor(str(row.get("seller") or row.get("vendor") or ""),str(row.get("product_link") or "")),"product":row.get("title") or row.get("product"),"model":row.get("model", ""),"quantity":qty,"unit_price":unit,"shipping":row.get("shipping",0),"estimated_tax":calc["tax"],"line_delivered_total":calc["delivered_total"],"product_link":row.get("product_link") or row.get("link", ""),"review_status":"DRAFT - VERIFY BEFORE ORDERING"})
    po=add_sheet("PO Draft",po_rows)
    for row in range(2,po.max_row+1):
        for col in range(1,po.max_column+1): po.cell(row,col).font=Font(color="0000FF" if po.cell(1,col).value in {"quantity","unit_price","shipping"} else "008000" if po.cell(1,col).value=="product_link" else "000000")
    for header in {"unit_price","shipping","estimated_tax","line_delivered_total"}:
        hm={c.value:c.column for c in po[1]}; col=hm.get(header)
        if col:
            for r in range(2,po.max_row+1): po.cell(r,col).number_format='$#,##0.00;[Red]($#,##0.00);-'
    buffer=BytesIO(); wb.save(buffer); buffer.seek(0)
    return safe_filename(f"{project_name or 'Procurement'}_Control_Center.xlsx"), buffer.getvalue()


def append_audit(audit_log: list[dict[str, Any]], action: str, item: str = "", details: str = "", user: str = "") -> None:
    audit_log.append({"timestamp":datetime.now().astimezone().isoformat(timespec="seconds"),"user":user,"action":action,"item":item,"details":details})
