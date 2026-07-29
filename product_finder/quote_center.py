from __future__ import annotations

from dataclasses import dataclass, asdict
from io import BytesIO
from typing import Any
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class QuoteLine:
    vendor: str = ""
    item_tag: str = ""
    manufacturer: str = ""
    model: str = ""
    description: str = ""
    quantity: float = 1.0
    unit_price: float = 0.0
    freight: float = 0.0
    tax: float = 0.0
    lead_time_days: float | None = None
    stock_status: str = ""
    substitution: str = ""
    quote_expires: str = ""
    source_file: str = ""

    @property
    def extended_price(self) -> float:
        return max(self.quantity, 0) * max(self.unit_price, 0)

    @property
    def landed_total(self) -> float:
        return self.extended_price + max(self.freight, 0) + max(self.tax, 0)

    def to_row(self) -> dict[str, Any]:
        row = asdict(self)
        row["extended_price"] = self.extended_price
        row["landed_total"] = self.landed_total
        return row


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _number(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    text = re.sub(r"[^0-9.()-]", "", str(value)).replace("(", "-").replace(")", "")
    try:
        return float(text)
    except (TypeError, ValueError):
        return default


ALIASES = {
    "vendor": ["vendor", "supplier", "seller", "company"],
    "item_tag": ["item tag", "item", "tag", "line item"],
    "manufacturer": ["manufacturer", "brand", "mfr"],
    "model": ["model", "model number", "mpn", "part number", "sku"],
    "description": ["description", "product", "title", "item description"],
    "quantity": ["quantity", "qty"],
    "unit_price": ["unit price", "price", "unit cost", "cost each"],
    "freight": ["freight", "shipping", "delivery charge"],
    "tax": ["tax", "sales tax"],
    "lead_time_days": ["lead time days", "lead time", "days"],
    "stock_status": ["stock status", "availability", "stock"],
    "substitution": ["substitution", "alternate", "exception"],
    "quote_expires": ["quote expires", "expiration", "valid through"],
}


def _map_columns(columns: list[str]) -> dict[str, str]:
    normalized = {_norm(c): c for c in columns}
    mapping: dict[str, str] = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[field] = normalized[alias]
                break
    return mapping


def parse_quote_file(data: bytes, filename: str, vendor_hint: str = "") -> list[QuoteLine]:
    lower = filename.lower()
    frames: list[pd.DataFrame] = []
    if lower.endswith(".csv"):
        frames = [pd.read_csv(BytesIO(data))]
    elif lower.endswith((".xlsx", ".xlsm")):
        book = pd.ExcelFile(BytesIO(data))
        for sheet in book.sheet_names:
            try:
                frame = pd.read_excel(book, sheet_name=sheet)
                if not frame.empty:
                    frames.append(frame)
            except Exception:
                continue
    else:
        raise ValueError("Quote Center supports CSV and XLSX files.")

    lines: list[QuoteLine] = []
    for frame in frames:
        frame = frame.dropna(how="all")
        if frame.empty:
            continue
        frame.columns = [str(c).strip() for c in frame.columns]
        mapping = _map_columns(list(frame.columns))
        if not any(k in mapping for k in ("description", "model", "unit_price")):
            continue
        for _, row in frame.iterrows():
            description = str(row.get(mapping.get("description", ""), "") or "").strip()
            model = str(row.get(mapping.get("model", ""), "") or "").strip()
            if not description and not model:
                continue
            vendor = str(row.get(mapping.get("vendor", ""), "") or vendor_hint or filename.rsplit(".", 1)[0]).strip()
            lines.append(QuoteLine(
                vendor=vendor,
                item_tag=str(row.get(mapping.get("item_tag", ""), "") or "").strip(),
                manufacturer=str(row.get(mapping.get("manufacturer", ""), "") or "").strip(),
                model=model,
                description=description,
                quantity=max(_number(row.get(mapping.get("quantity", ""), 1), 1), 0),
                unit_price=max(_number(row.get(mapping.get("unit_price", ""), 0)), 0),
                freight=max(_number(row.get(mapping.get("freight", ""), 0)), 0),
                tax=max(_number(row.get(mapping.get("tax", ""), 0)), 0),
                lead_time_days=_number(row.get(mapping.get("lead_time_days", ""), ""), 0) or None,
                stock_status=str(row.get(mapping.get("stock_status", ""), "") or "").strip(),
                substitution=str(row.get(mapping.get("substitution", ""), "") or "").strip(),
                quote_expires=str(row.get(mapping.get("quote_expires", ""), "") or "").strip(),
                source_file=filename,
            ))
    return lines


def quote_dataframe(lines: list[QuoteLine]) -> pd.DataFrame:
    columns = list(QuoteLine().__dict__.keys()) + ["extended_price", "landed_total"]
    if not lines:
        return pd.DataFrame(columns=columns)
    df = pd.DataFrame([line.to_row() for line in lines])
    for col in ["quantity", "unit_price", "freight", "tax", "lead_time_days", "extended_price", "landed_total"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    return df


def vendor_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["vendor", "line_count", "quoted_total", "avg_lead_days", "substitution_count", "coverage_score"])
    work = df.copy()
    work["has_substitution"] = work["substitution"].fillna("").astype(str).str.strip().ne("")
    summary = work.groupby("vendor", dropna=False).agg(
        line_count=("vendor", "size"),
        quoted_total=("landed_total", "sum"),
        avg_lead_days=("lead_time_days", "mean"),
        substitution_count=("has_substitution", "sum"),
        unique_items=("model", lambda s: s.fillna("").astype(str).replace("", pd.NA).nunique()),
    ).reset_index()
    max_items = max(int(summary["unique_items"].max() or 1), 1)
    summary["coverage_score"] = (summary["unique_items"] / max_items * 100).round(1)
    summary["avg_lead_days"] = summary["avg_lead_days"].fillna(0).round(1)
    summary = summary.sort_values(["coverage_score", "quoted_total"], ascending=[False, True]).reset_index(drop=True)
    return summary[["vendor", "line_count", "quoted_total", "avg_lead_days", "substitution_count", "coverage_score"]]


def award_recommendation(summary: pd.DataFrame) -> dict[str, Any]:
    if summary.empty:
        return {}
    work = summary.copy()
    max_total = max(float(work["quoted_total"].max()), 1.0)
    max_lead = max(float(work["avg_lead_days"].max()), 1.0)
    work["score"] = (
        work["coverage_score"] * 0.50
        + (1 - work["quoted_total"] / max_total) * 100 * 0.30
        + (1 - work["avg_lead_days"] / max_lead) * 100 * 0.15
        + (1 - (work["substitution_count"] / work["line_count"].clip(lower=1))) * 100 * 0.05
    ).round(1)
    best = work.sort_values(["score", "quoted_total"], ascending=[False, True]).iloc[0].to_dict()
    return best


def create_bid_tab_workbook(lines: list[QuoteLine], project_name: str = "") -> bytes:
    df = quote_dataframe(lines)
    summary = vendor_summary(df)
    recommendation = award_recommendation(summary)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bid Summary"
    navy, blue, light, green = "172B4D", "0F6CBD", "EAF3FB", "E2F0D9"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"BID TAB - {project_name or 'PROCUREMENT PROJECT'}"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["Vendor", "Lines", "Quoted Total", "Avg Lead Days", "Substitutions", "Coverage"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=blue)
    for r_idx, row in enumerate(summary.to_dict("records"), 4):
        vals = [row["vendor"], row["line_count"], row["quoted_total"], row["avg_lead_days"], row["substitution_count"], row["coverage_score"] / 100]
        for c, value in enumerate(vals, 1):
            ws.cell(r_idx, c, value)
        ws.cell(r_idx, 3).number_format = '$#,##0.00'
        ws.cell(r_idx, 6).number_format = '0.0%'
        if recommendation and row["vendor"] == recommendation.get("vendor"):
            for c in range(1, 7): ws.cell(r_idx, c).fill = PatternFill("solid", fgColor=green)
    for i, width in enumerate([24, 10, 18, 16, 16, 14], 1): ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"; ws.sheet_view.showGridLines = False

    detail = wb.create_sheet("Quote Lines")
    headers2 = list(df.columns)
    for c, h in enumerate(headers2, 1):
        cell = detail.cell(1, c, h.replace("_", " ").title()); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=blue)
    for r_idx, row in enumerate(df.to_dict("records"), 2):
        for c, h in enumerate(headers2, 1): detail.cell(r_idx, c, row.get(h, ""))
    for col in ["unit_price", "freight", "tax", "extended_price", "landed_total"]:
        if col in headers2:
            c = headers2.index(col) + 1
            for r in range(2, len(df) + 2): detail.cell(r, c).number_format = '$#,##0.00'
    detail.freeze_panes = "A2"; detail.auto_filter.ref = detail.dimensions; detail.sheet_view.showGridLines = False
    for c in range(1, len(headers2) + 1): detail.column_dimensions[get_column_letter(c)].width = 18

    bio = BytesIO(); wb.save(bio); return bio.getvalue()
