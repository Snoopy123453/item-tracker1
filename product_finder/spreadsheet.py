from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
from typing import Any, Iterable

import requests
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.worksheet import Worksheet

from .models import InputRecord, ProductResult, SpecDocument, StoreResult
from .utils import ensure_directory, safe_filename

NAVY = "17324D"
BLUE = "2E75B6"
LIGHT_BLUE = "DCEAF7"
PALE_BLUE = "EFF6FC"
GREEN = "008000"
GRAY = "666666"
WHITE = "FFFFFF"
ORANGE = "FCE4D6"
TEAL = "DDEBF7"
PURPLE = "7030A0"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color=WHITE, bold=True)
TITLE_FONT = Font(size=20, bold=True, color=NAVY)
SUBTITLE_FONT = Font(size=11, italic=True, color=GRAY)
IMPORTED_FONT = Font(color=GREEN)
STATIC_FONT = Font(color=GRAY)
CAUTION_FILL = PatternFill("solid", fgColor=ORANGE)
TEAL_FILL = PatternFill("solid", fgColor=TEAL)
THIN_BLUE = Side(style="thin", color="9EADCC")
INVALID_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
USER_AGENT = "ProductHunterWebApp/3.0"


def _excel_safe(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    cleaned = INVALID_XML_CHARS.sub("", value)
    if cleaned.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + cleaned
    return cleaned


def _write_rows(ws: Worksheet, headers: list[str], rows: Iterable[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([_excel_safe(row.get(header, "")) for header in headers])


def _format_table(ws: Worksheet, *, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_BLUE)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = PatternFill("solid", fgColor=PALE_BLUE) if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = fill
            if isinstance(cell.value, str) and cell.value.startswith(("https://", "http://")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
            else:
                cell.font = IMPORTED_FONT
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        max_len = min(max([len(value) for value in values] + [10]), 52)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)
    ws.row_dimensions[1].height = 34


def _format_number_columns(ws: Worksheet, currency_headers: set[str], number_headers: set[str]) -> None:
    header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
    for header in currency_headers:
        col = header_map.get(header)
        if col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = '$#,##0.00;[Red]($#,##0.00);-'
    for header in number_headers:
        col = header_map.get(header)
        if col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = '#,##0.0'


def _download_thumbnail(url: str, timeout: int = 8) -> BytesIO | None:
    if not url.startswith(("https://", "http://")):
        return None
    try:
        response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=(4, timeout))
        response.raise_for_status()
        if len(response.content) > 4_000_000:
            return None
        image = PILImage.open(BytesIO(response.content)).convert("RGB")
        image.thumbnail((180, 120))
        output = BytesIO()
        image.save(output, format="PNG", optimize=True)
        output.seek(0)
        return output
    except Exception:
        return None


def _embed_product_images(ws: Worksheet, product_results: list[ProductResult]) -> int:
    if not product_results:
        return 0
    header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
    image_col = header_map.get("product_image")
    if not image_col:
        return 0
    ws.column_dimensions[get_column_letter(image_col)].width = 24
    count = 0
    for row_idx, result in enumerate(product_results, start=2):
        stream = _download_thumbnail(result.thumbnail)
        ws.row_dimensions[row_idx].height = 92
        if stream is None:
            ws.cell(row=row_idx, column=image_col, value="Image unavailable")
            ws.cell(row=row_idx, column=image_col).font = STATIC_FONT
            continue
        image = XLImage(stream)
        image.width = 120
        image.height = 80
        image.anchor = ws.cell(row=row_idx, column=image_col).coordinate
        ws.add_image(image)
        count += 1
    return count


def _report_subjects(input_records: list[InputRecord]) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for record in input_records:
        candidate = record.extracted_product_name or (record.label if record.input_type == "text" else "")
        candidate = re.sub(r"\s+", " ", candidate).strip(" ._-")
        if not candidate:
            continue
        key = candidate.casefold()
        if key not in seen:
            seen.add(key)
            names.append(candidate)
    return names


def suggest_workbook_filename(input_records: list[InputRecord]) -> str:
    subjects = _report_subjects(input_records)
    if not subjects:
        stem = "Product_Search_Report"
    elif len(subjects) == 1:
        stem = subjects[0]
    elif len(subjects) == 2:
        stem = f"{subjects[0]}_and_{subjects[1]}"
    else:
        shared = " ".join(subjects).lower()
        if any(term in shared for term in ("sink", "drain", "faucet", "plumbing", "trap primer", "eyewash")):
            stem = "Plumbing_Fixture_Procurement_Report"
        elif any(term in shared for term in ("panel", "breaker", "transformer", "electrical")):
            stem = "Electrical_Equipment_Procurement_Report"
        else:
            stem = f"Product_Procurement_Report_{len(subjects)}_Items"
    return safe_filename(f"{stem[:120]}.xlsx")


def _add_best_matches_sheet(wb: Workbook, product_results: list[ProductResult]) -> Worksheet:
    ws = wb.create_sheet("Best Matches")
    headers = ["query", "match_score", "match_grade", "match_confidence", "match_profile", "exact_model_match", "title", "seller", "price", "product_link", "matched_features", "missing_features", "differences", "score_breakdown", "recommendation"]
    rows = []
    for result in product_results:
        if result.best_match:
            row = result.to_row()
            rows.append(row)
    _write_rows(ws, headers, rows)
    if not rows:
        ws.append(["No product listings returned"] + [""] * 14)
    _format_table(ws)
    _format_number_columns(ws, set(), {"match_score"})
    header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
    score_col = header_map.get("match_score")
    if score_col and ws.max_row >= 2:
        letter = get_column_letter(score_col)
        ws.conditional_formatting.add(f"{letter}2:{letter}{ws.max_row}", CellIsRule(operator="greaterThanOrEqual", formula=["90"], fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add(f"{letter}2:{letter}{ws.max_row}", CellIsRule(operator="between", formula=["70", "89.999"], fill=PatternFill("solid", fgColor="FFEB9C")))
        ws.conditional_formatting.add(f"{letter}2:{letter}{ws.max_row}", CellIsRule(operator="lessThan", formula=["70"], fill=PatternFill("solid", fgColor="FFC7CE")))
    return ws


def _add_comparison_sheet(wb: Workbook, product_results: list[ProductResult]) -> Worksheet:
    ws = wb.create_sheet("Price Comparison")
    headers = ["query", "listing_count", "lowest_price", "average_price", "highest_price", "lowest_price_seller", "review_status"]
    ws.append(headers)
    queries: list[str] = []
    seen: set[str] = set()
    for result in product_results:
        key = result.query.casefold()
        if key not in seen:
            seen.add(key)
            queries.append(result.query)
    product_sheet = quote_sheetname("Product Results")
    last_row = max(2, len(product_results) + 1)
    for row_idx, query in enumerate(queries, start=2):
        ws.cell(row_idx, 1, _excel_safe(query))
        ws.cell(row_idx, 2, f'=COUNTIF({product_sheet}!$B$2:$B${last_row},A{row_idx})')
        ws.cell(row_idx, 3, f'=IFERROR(MINIFS({product_sheet}!$H$2:$H${last_row},{product_sheet}!$B$2:$B${last_row},A{row_idx},{product_sheet}!$H$2:$H${last_row},">0"),"")')
        ws.cell(row_idx, 4, f'=IFERROR(AVERAGEIFS({product_sheet}!$H$2:$H${last_row},{product_sheet}!$B$2:$B${last_row},A{row_idx},{product_sheet}!$H$2:$H${last_row},">0"),"")')
        ws.cell(row_idx, 5, f'=IFERROR(MAXIFS({product_sheet}!$H$2:$H${last_row},{product_sheet}!$B$2:$B${last_row},A{row_idx}),"")')
        ws.cell(row_idx, 6, "Open Product Results and filter by query")
        ws.cell(row_idx, 7, '=IF(B%d=0,"No listings",IF(B%d<3,"Limited comparison","Review retailer terms"))' % (row_idx, row_idx))
    if not queries:
        ws.append(["No priced product listings returned", 0, "", "", "", "", "No comparison available"])
    _format_table(ws)
    _format_number_columns(ws, {"lowest_price", "average_price", "highest_price"}, {"listing_count"})
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["F"].width = 34
    return ws


def _add_notes_sheet(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Review Notes")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "PROCUREMENT REVIEW CHECKLIST"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    headers = ["item / model", "reviewer", "status", "decision / notes", "date reviewed"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(3, col, value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    statuses = ["Confirm exact model", "Confirm dimensions", "Confirm voltage/connection", "Confirm code compliance", "Confirm stock and lead time", "Confirm warranty", "Approve substitution only after review"]
    for row, status in enumerate(statuses, start=4):
        ws.cell(row, 1, status)
        ws.cell(row, 3, "Pending")
        ws.cell(row, 3).fill = CAUTION_FILL
    widths = [34, 20, 18, 55, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A4"
    return ws


def _add_summary_sheet(wb: Workbook, *, product_count: int, store_count: int, spec_count: int, image_count: int, input_count: int, location: str, run_notes: str, report_title: str) -> Worksheet:
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = report_title.upper()
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.row_dimensions[1].height = 38
    ws.merge_cells("A2:F2")
    ws["A2"] = "Ranked best matches, retail listings, nearby suppliers, embedded product images, price comparisons, and technical documents."
    ws["A2"].font = SUBTITLE_FONT

    cards = [
        ("A4", "Inputs", input_count), ("C4", "Listings", product_count), ("E4", "Nearby stores", store_count),
        ("A7", "Documents", spec_count), ("C7", "Images embedded", image_count), ("E7", "Search location", location or "Not specified"),
    ]
    for cell_ref, label, value in cards:
        col, row = ws[cell_ref].column, ws[cell_ref].row
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        value_cell = ws.cell(row=row + 1, column=col, value=_excel_safe(value))
        value_cell.font = Font(bold=True, color=NAVY, size=16)
        value_cell.fill = TEAL_FILL
        value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A11:F11")
    ws["A11"] = "Run notes"
    ws["A11"].font = HEADER_FONT
    ws["A11"].fill = HEADER_FILL
    ws.merge_cells("A12:F13")
    ws["A12"] = _excel_safe(run_notes or "No warnings were recorded.")
    ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A12"].fill = PatternFill("solid", fgColor=PALE_BLUE)

    ws.merge_cells("A15:F15")
    ws["A15"] = "Verification required"
    ws["A15"].font = Font(bold=True, color="9C6500")
    ws["A15"].fill = CAUTION_FILL
    ws.merge_cells("A16:F19")
    ws["A16"] = (
        "Prices, shipping, stock, and availability change. Nearby-store results are leads rather than guaranteed shelf inventory. "
        "Verify manufacturer, exact model number, dimensions, ratings, code requirements, warranty, and document revision before ordering. "
        "Replacement or equivalent products require professional review and approval."
    )
    ws["A16"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A16"].fill = CAUTION_FILL
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 19
    return ws


def _build_workbook(*, input_records: list[InputRecord], product_results: list[ProductResult], store_results: list[StoreResult], spec_documents: list[SpecDocument], location: str, run_notes: str = "") -> Workbook:
    wb = Workbook()
    subjects = _report_subjects(input_records)
    report_title = subjects[0] if len(subjects) == 1 else ("Product Procurement Report" if not subjects else f"Product Procurement Report — {len(subjects)} Items")
    wb.properties.title = report_title
    wb.properties.subject = "Retail products, nearby stores, product images, price comparisons, and technical documents"
    wb.properties.creator = "Product Hunter Web App"

    input_headers = ["input_type", "label", "extracted_product_name", "brand", "category", "confidence", "generated_queries", "notes", "source_url", "retrieved_at"]
    ws_inputs = wb.create_sheet("Inputs")
    _write_rows(ws_inputs, input_headers, [record.to_row() for record in input_records])
    _format_table(ws_inputs)
    _format_number_columns(ws_inputs, set(), {"confidence"})

    product_headers = ["product_image", "query", "input_source", "rank", "title", "seller", "price", "extracted_price", "delivery", "rating", "reviews", "condition", "snippet", "product_link", "seller_link", "thumbnail", "search_location", "retrieved_at", "raw_source", "best_match", "match_score", "match_grade", "match_confidence", "match_profile", "exact_model_match", "matched_features", "missing_features", "differences", "score_breakdown", "recommendation"]
    product_rows = []
    for result in product_results:
        row = result.to_row(); row["product_image"] = ""; product_rows.append(row)
    ws_products = wb.create_sheet("Product Results")
    _write_rows(ws_products, product_headers, product_rows)
    _format_table(ws_products)
    _format_number_columns(ws_products, {"extracted_price"}, {"rating", "reviews", "match_score"})
    product_header_map = {cell.value: cell.column for cell in ws_products[1] if cell.value}
    match_col = product_header_map.get("match_score")
    best_col = product_header_map.get("best_match")
    if match_col and ws_products.max_row >= 2:
        letter = get_column_letter(match_col)
        ws_products.conditional_formatting.add(f"{letter}2:{letter}{ws_products.max_row}", CellIsRule(operator="greaterThanOrEqual", formula=["90"], fill=PatternFill("solid", fgColor="C6EFCE")))
        ws_products.conditional_formatting.add(f"{letter}2:{letter}{ws_products.max_row}", CellIsRule(operator="between", formula=["70", "89.999"], fill=PatternFill("solid", fgColor="FFEB9C")))
        ws_products.conditional_formatting.add(f"{letter}2:{letter}{ws_products.max_row}", CellIsRule(operator="lessThan", formula=["70"], fill=PatternFill("solid", fgColor="FFC7CE")))
    if best_col and ws_products.max_row >= 2:
        best_letter = get_column_letter(best_col)
        ws_products.conditional_formatting.add(f"A2:AA{ws_products.max_row}", FormulaRule(formula=[f'${best_letter}2=TRUE'], fill=PatternFill("solid", fgColor="E2F0D9")))
    image_count = _embed_product_images(ws_products, product_results)

    store_headers = ["query", "rank", "title", "store_type", "address", "phone", "rating", "reviews", "hours", "website", "directions", "maps_link", "search_location", "retrieved_at", "raw_source"]
    ws_stores = wb.create_sheet("Nearby Stores")
    _write_rows(ws_stores, store_headers, [result.to_row() for result in store_results])
    _format_table(ws_stores)
    _format_number_columns(ws_stores, set(), {"rating", "reviews"})

    spec_headers = ["query", "rank", "title", "document_type", "source_domain", "link", "displayed_link", "snippet", "official_source", "pdf_link", "match_confidence", "retrieved_at", "raw_source"]
    ws_specs = wb.create_sheet("Spec Documents")
    _write_rows(ws_specs, spec_headers, [result.to_row() for result in spec_documents])
    _format_table(ws_specs)

    _add_best_matches_sheet(wb, product_results)
    _add_comparison_sheet(wb, product_results)
    _add_notes_sheet(wb)
    dashboard = _add_summary_sheet(wb, product_count=len(product_results), store_count=len(store_results), spec_count=len(spec_documents), image_count=image_count, input_count=len(input_records), location=location, run_notes=run_notes, report_title=report_title)
    wb._sheets.remove(dashboard); wb._sheets.insert(0, dashboard); wb.active = 0
    return wb


def create_product_workbook_bytes(*, input_records: list[InputRecord], product_results: list[ProductResult], store_results: list[StoreResult], spec_documents: list[SpecDocument] | None = None, location: str, run_notes: str = "") -> tuple[str, bytes]:
    filename = suggest_workbook_filename(input_records)
    wb = _build_workbook(input_records=input_records, product_results=product_results, store_results=store_results, spec_documents=spec_documents or [], location=location, run_notes=run_notes)
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return filename, buffer.getvalue()


def create_product_workbook(*, input_records: list[InputRecord], product_results: list[ProductResult], store_results: list[StoreResult], spec_documents: list[SpecDocument] | None = None, location: str, output_dir: str | Path = "exports", run_notes: str = "") -> Path:
    output_path = ensure_directory(output_dir)
    filename, workbook_bytes = create_product_workbook_bytes(input_records=input_records, product_results=product_results, store_results=store_results, spec_documents=spec_documents or [], location=location, run_notes=run_notes)
    workbook_path = output_path / filename; workbook_path.write_bytes(workbook_bytes)
    return workbook_path
