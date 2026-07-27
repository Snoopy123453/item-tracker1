from __future__ import annotations

import base64
import json
import re
import zipfile
from dataclasses import dataclass, asdict
from datetime import date
from io import BytesIO
from typing import Any

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from .utils import clean_text, extract_json_object, safe_filename


@dataclass
class EquipmentItem:
    item_tag: str = ""
    division: str = ""
    manufacturer: str = ""
    model: str = ""
    description: str = ""
    quantity: int = 1
    location: str = ""
    source_file: str = ""
    source_page: str = ""
    status: str = "Needs search"
    approved_listing: str = ""
    notes: str = ""

    def to_row(self) -> dict[str, Any]:
        return asdict(self)


EXTRACTION_PROMPT = """
Extract purchasable equipment/products from the supplied schedule text or image.
Return ONLY JSON with shape:
{"items":[{"item_tag":"","division":"","manufacturer":"","model":"","description":"","quantity":1,"location":"","source_page":"","notes":""}]}
Rules:
- One row per distinct purchasable model or component.
- Preserve exact model numbers, suffixes, dimensions, ratings, materials, and accessories.
- Never invent a manufacturer/model. Leave blank when not stated.
- quantity must be an integer and defaults to 1.
- division may be Plumbing, Mechanical, Electrical, Architectural, Equipment, or General.
""".strip()


def _items_from_payload(payload: dict[str, Any], source_file: str) -> list[EquipmentItem]:
    output: list[EquipmentItem] = []
    items = payload.get("items", []) if isinstance(payload, dict) else []
    if not isinstance(items, list):
        return output
    for raw in items:
        if not isinstance(raw, dict):
            continue
        try:
            qty = max(1, int(raw.get("quantity") or 1))
        except (TypeError, ValueError):
            qty = 1
        output.append(EquipmentItem(
            item_tag=clean_text(raw.get("item_tag")),
            division=clean_text(raw.get("division")) or "General",
            manufacturer=clean_text(raw.get("manufacturer")),
            model=clean_text(raw.get("model")),
            description=clean_text(raw.get("description")),
            quantity=qty,
            location=clean_text(raw.get("location")),
            source_file=source_file,
            source_page=clean_text(raw.get("source_page")),
            notes=clean_text(raw.get("notes")),
        ))
    return output


def extract_pdf_text(data: bytes, max_pages: int = 80) -> str:
    reader = PdfReader(BytesIO(data))
    chunks: list[str] = []
    for idx, page in enumerate(reader.pages[:max_pages], start=1):
        text = page.extract_text() or ""
        if text.strip():
            chunks.append(f"\n--- PAGE {idx} ---\n{text}")
    return "".join(chunks)[:180000]


def extract_schedule_items(*, file_bytes: bytes, filename: str, mime_type: str, openai_api_key: str, model: str) -> list[EquipmentItem]:
    if not openai_api_key:
        raise ValueError("OpenAI API key is required for schedule extraction.")
    from openai import OpenAI
    client = OpenAI(api_key=openai_api_key)
    lower = filename.lower()
    if lower.endswith(".pdf") or mime_type == "application/pdf":
        text = extract_pdf_text(file_bytes)
        if not text.strip():
            raise ValueError("No selectable text was found in this PDF. Export schedule pages as images and upload them instead.")
        content = [{"type": "input_text", "text": EXTRACTION_PROMPT + "\n\nDOCUMENT TEXT:\n" + text}]
    elif lower.endswith((".txt", ".csv")) or mime_type.startswith("text/"):
        text = file_bytes.decode("utf-8", errors="replace")[:180000]
        content = [{"type": "input_text", "text": EXTRACTION_PROMPT + "\n\nDOCUMENT TEXT:\n" + text}]
    else:
        b64 = base64.b64encode(file_bytes).decode("ascii")
        media = mime_type if mime_type.startswith("image/") else "image/jpeg"
        content = [
            {"type": "input_text", "text": EXTRACTION_PROMPT},
            {"type": "input_image", "image_url": f"data:{media};base64,{b64}"},
        ]
    response = client.responses.create(model=model, input=[{"role": "user", "content": content}])
    payload = extract_json_object(getattr(response, "output_text", "") or "")
    return _items_from_payload(payload, filename)


def consolidate_items(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for row in rows:
        manufacturer = clean_text(row.get("manufacturer"))
        model = clean_text(row.get("model"))
        description = clean_text(row.get("description"))
        key = (manufacturer + "|" + model).casefold() if model else re.sub(r"\W+", " ", description.casefold()).strip()
        key = key or f"row-{len(merged)}"
        qty = max(1, int(row.get("quantity") or 1))
        if key not in merged:
            merged[key] = dict(row)
            merged[key]["quantity"] = qty
            merged[key]["item_tag"] = clean_text(row.get("item_tag"))
            merged[key]["location"] = clean_text(row.get("location"))
            merged[key]["source_file"] = clean_text(row.get("source_file"))
        else:
            current = merged[key]
            current["quantity"] = int(current.get("quantity") or 0) + qty
            for field in ("item_tag", "location", "source_file", "source_page"):
                values = [v for v in [clean_text(current.get(field)), clean_text(row.get(field))] if v]
                current[field] = "; ".join(dict.fromkeys("; ".join(values).split("; ")))
    return list(merged.values())


def create_project_backup(project: dict[str, Any]) -> tuple[str, bytes]:
    payload = json.dumps(project, indent=2, ensure_ascii=False).encode("utf-8")
    name = safe_filename(f"{project.get('project_name') or 'Product_Hunter_Project'}_backup.json")
    return name, payload


def load_project_backup(data: bytes) -> dict[str, Any]:
    payload = json.loads(data.decode("utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("equipment", []), list):
        raise ValueError("This is not a valid Product Hunter project backup.")
    return payload


def create_project_workbook(project: dict[str, Any]) -> tuple[str, bytes]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Dashboard"
    navy = "17324D"; blue = "2E75B6"; pale = "EAF3FB"; teal = "DDEBF7"; white = "FFFFFF"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1"); ws["A1"] = str(project.get("project_name") or "PROJECT PROCUREMENT").upper()
    ws["A1"].font = Font(size=20, bold=True, color=white); ws["A1"].fill = PatternFill("solid", fgColor=navy)
    equipment = project.get("equipment", [])
    metrics = [("Products", len(equipment)), ("Total quantity", sum(int(r.get("quantity") or 0) for r in equipment)), ("Approved", sum(str(r.get("status", "")).lower()=="approved" for r in equipment)), ("Needs review", sum("review" in str(r.get("status", "")).lower() for r in equipment))]
    for i, (label, value) in enumerate(metrics):
        col = 1 + (i % 2) * 3; row = 3 + (i // 2) * 3
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        ws.cell(row, col, label).fill = PatternFill("solid", fgColor=blue); ws.cell(row,col).font=Font(bold=True,color=white)
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        ws.cell(row+1,col,value).fill=PatternFill("solid",fgColor=teal); ws.cell(row+1,col).font=Font(size=16,bold=True,color=navy)
    for c in range(1,7): ws.column_dimensions[get_column_letter(c)].width=20

    eq = wb.create_sheet("Equipment Register")
    headers = ["item_tag","division","manufacturer","model","description","quantity","location","source_file","source_page","status","approved_listing","notes"]
    eq.append(headers)
    for row in equipment: eq.append([row.get(h, "") for h in headers])
    eq.freeze_panes="A2"; eq.auto_filter.ref=eq.dimensions; eq.sheet_view.showGridLines=False
    for cell in eq[1]: cell.fill=PatternFill("solid",fgColor=navy); cell.font=Font(bold=True,color=white); cell.alignment=Alignment(wrap_text=True)
    for idx, col in enumerate(headers,1): eq.column_dimensions[get_column_letter(idx)].width = 16 if col not in {"description","notes","approved_listing"} else 42
    for row in eq.iter_rows(min_row=2):
        for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)
        if row[10].value and str(row[10].value).startswith("http"):
            row[10].hyperlink=str(row[10].value); row[10].style="Hyperlink"

    prefs = wb.create_sheet("Project Rules")
    prefs.append(["Setting","Value"])
    for k,v in (project.get("preferences") or {}).items(): prefs.append([k, str(v)])
    for cell in prefs[1]: cell.fill=PatternFill("solid",fgColor=navy); cell.font=Font(bold=True,color=white)
    prefs.column_dimensions["A"].width=36; prefs.column_dimensions["B"].width=60

    buffer=BytesIO(); wb.save(buffer); buffer.seek(0)
    return safe_filename(f"{project.get('project_name') or 'Project'}_Procurement.xlsx"), buffer.getvalue()


def create_submittal_zip(project: dict[str, Any], document_rows: list[dict[str, Any]]) -> tuple[str, bytes]:
    output = BytesIO()
    manifest = ["PRODUCT HUNTER SUBMITTAL PACKAGE", f"Project: {project.get('project_name','')}", f"Created: {date.today().isoformat()}", ""]
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, row in enumerate(document_rows, start=1):
            url = clean_text(row.get("link") or row.get("product_link"))
            title = clean_text(row.get("title") or row.get("product") or f"Document {idx}")
            manifest.append(f"{idx}. {title}\n   {url}")
            if url.lower().endswith(".pdf"):
                try:
                    response = requests.get(url, timeout=(8, 25), headers={"User-Agent":"ProductHunter/1.0"})
                    if response.ok and "pdf" in response.headers.get("content-type", "").lower() and len(response.content) < 30_000_000:
                        zf.writestr(safe_filename(f"{idx:02d}_{title[:70]}.pdf"), response.content)
                except requests.RequestException:
                    pass
        zf.writestr("MANIFEST.txt", "\n".join(manifest))
        project_name, project_bytes = create_project_workbook(project)
        zf.writestr(project_name, project_bytes)
    output.seek(0)
    return safe_filename(f"{project.get('project_name') or 'Project'}_Submittal_Package.zip"), output.getvalue()
