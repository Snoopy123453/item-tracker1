from __future__ import annotations

import base64
import json
import re
from dataclasses import dataclass, asdict
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from PIL import Image, ImageOps

from .utils import clean_text, extract_json_object


@dataclass
class SpecAttribute:
    category: str
    attribute: str
    value: str
    unit: str = ""
    requirement_level: str = "Required"
    source_page: str = ""
    evidence: str = ""


@dataclass
class AttributeComparison:
    category: str
    attribute: str
    original_value: str
    candidate_value: str
    original_unit: str
    candidate_unit: str
    requirement_level: str
    status: str
    confidence: float
    explanation: str
    original_page: str = ""
    candidate_page: str = ""


@dataclass
class CandidateComparison:
    candidate_name: str
    manufacturer: str
    model: str
    status: str
    compatibility_score: float
    evidence_coverage: float
    hard_conflicts: int
    unconfirmed_required: int
    comparisons: list[AttributeComparison]
    summary: str


EXTRACT_PROMPT = """
You extract procurement specifications from a technical product document.
Return ONLY JSON with this exact structure:
{
  "document_title": "",
  "manufacturer": "",
  "model": "",
  "product_type": "",
  "attributes": [
    {
      "category": "Identity|Dimensions|Connections|Performance|Materials|Finish|Electrical|Compliance|Accessories|Installation|Other",
      "attribute": "concise normalized attribute name",
      "value": "exact value from document",
      "unit": "unit if present, otherwise empty",
      "requirement_level": "Required|Preferred|Optional",
      "source_page": "page number or empty",
      "evidence": "short exact supporting phrase"
    }
  ],
  "notes": ""
}
Rules:
- Never invent specifications.
- Preserve model suffixes, sizes, ratings, standards, and included accessories exactly.
- Split combined statements into separate attributes.
- Treat explicit product characteristics as Required unless the document says optional.
- Keep evidence short.
""".strip()

COMPARE_PROMPT = """
You are a technical procurement verifier. Compare an original required specification against one candidate specification.
Return ONLY JSON:
{
  "comparisons": [
    {
      "attribute": "attribute name from original",
      "candidate_value": "candidate value or empty",
      "status": "Match|Equivalent|Conflict|Not Confirmed|Not Applicable",
      "confidence": 0.0,
      "explanation": "brief factual explanation",
      "candidate_page": "page or empty"
    }
  ],
  "summary": "brief overall technical conclusion"
}
Rules:
- Missing information is Not Confirmed, never Match.
- A different model number is a Conflict when exact model is required, otherwise note it without assuming incompatibility.
- Normalize equivalent units and terminology before deciding.
- Do not approve substitutions based only on similar product names.
- Any explicit mismatch in voltage, connection size/type, dimensions, material, pressure/flow/capacity, certification, mounting, finish, or required accessory is Conflict.
""".strip()


def _pdf_text(data: bytes, max_pages: int = 40) -> tuple[str, int]:
    reader = PdfReader(BytesIO(data))
    chunks: list[str] = []
    for i, page in enumerate(reader.pages[:max_pages], start=1):
        text = clean_text(page.extract_text() or "")
        if text:
            chunks.append(f"\n--- PAGE {i} ---\n{text}")
    return "".join(chunks), len(reader.pages)


def _image_data_url(data: bytes) -> str:
    with Image.open(BytesIO(data)) as original:
        image = ImageOps.exif_transpose(original).convert("RGB")
        image.thumbnail((2200, 2200))
        out = BytesIO()
        image.save(out, format="JPEG", quality=88, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(out.getvalue()).decode("ascii")


def extract_spec_document(*, data: bytes, filename: str, openai_api_key: str, model: str) -> dict[str, Any]:
    if not openai_api_key:
        raise ValueError("OpenAI API key is required for specification extraction.")
    suffix = Path(filename).suffix.lower()
    from openai import OpenAI
    client = OpenAI(api_key=openai_api_key)

    if suffix == ".pdf":
        text, page_count = _pdf_text(data)
        if not text.strip():
            raise ValueError("This PDF appears to be scanned and has no selectable text. Upload page images or a text-searchable PDF.")
        # Keep request size bounded while retaining page markers.
        text = text[:120000]
        content: list[dict[str, Any]] = [{"type": "input_text", "text": EXTRACT_PROMPT + f"\n\nFilename: {filename}\nPages: {page_count}\n\nDOCUMENT TEXT:\n{text}"}]
    elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        content = [
            {"type": "input_text", "text": EXTRACT_PROMPT + f"\n\nFilename: {filename}"},
            {"type": "input_image", "image_url": _image_data_url(data)},
        ]
    else:
        raise ValueError("Supported spec-sheet formats are PDF, PNG, JPG, JPEG, and WebP.")

    response = client.responses.create(model=model, input=[{"role": "user", "content": content}])
    parsed = extract_json_object(getattr(response, "output_text", "") or "")
    attrs: list[dict[str, Any]] = []
    for raw in parsed.get("attributes", []) if isinstance(parsed.get("attributes"), list) else []:
        if not isinstance(raw, dict):
            continue
        attribute = clean_text(raw.get("attribute"))
        value = clean_text(raw.get("value"))
        if not attribute or not value:
            continue
        attrs.append(asdict(SpecAttribute(
            category=clean_text(raw.get("category")) or "Other",
            attribute=attribute,
            value=value,
            unit=clean_text(raw.get("unit")),
            requirement_level=clean_text(raw.get("requirement_level")) or "Required",
            source_page=clean_text(raw.get("source_page")),
            evidence=clean_text(raw.get("evidence")),
        )))
    parsed["attributes"] = attrs
    parsed["filename"] = filename
    return parsed


def _norm(value: str) -> str:
    value = value.lower().replace("½", "1/2").replace("¼", "1/4").replace("¾", "3/4")
    value = re.sub(r"\s+", " ", value)
    return re.sub(r"[^a-z0-9./-]+", " ", value).strip()


def _deterministic_status(original: str, candidate: str) -> str | None:
    a, b = _norm(original), _norm(candidate)
    if not b:
        return "Not Confirmed"
    if a == b or a in b or b in a:
        return "Match"
    # Strong numeric contradiction signal.
    nums_a = re.findall(r"\d+(?:\.\d+)?(?:/\d+)?", a)
    nums_b = re.findall(r"\d+(?:\.\d+)?(?:/\d+)?", b)
    if nums_a and nums_b and set(nums_a) != set(nums_b):
        return "Conflict"
    return None


def compare_spec_documents(*, original: dict[str, Any], candidate: dict[str, Any], openai_api_key: str, model: str) -> CandidateComparison:
    original_attrs = original.get("attributes", [])
    candidate_attrs = candidate.get("attributes", [])
    prompt_payload = {
        "original": {
            "manufacturer": original.get("manufacturer", ""),
            "model": original.get("model", ""),
            "attributes": original_attrs,
        },
        "candidate": {
            "manufacturer": candidate.get("manufacturer", ""),
            "model": candidate.get("model", ""),
            "attributes": candidate_attrs,
        },
    }
    from openai import OpenAI
    client = OpenAI(api_key=openai_api_key)
    response = client.responses.create(
        model=model,
        input=COMPARE_PROMPT + "\n\nDATA:\n" + json.dumps(prompt_payload, ensure_ascii=False),
    )
    ai = extract_json_object(getattr(response, "output_text", "") or "")
    by_name = {clean_text(x.get("attribute")).lower(): x for x in ai.get("comparisons", []) if isinstance(x, dict)}
    candidate_by_name = {clean_text(x.get("attribute")).lower(): x for x in candidate_attrs if isinstance(x, dict)}

    comparisons: list[AttributeComparison] = []
    for orig in original_attrs:
        name = clean_text(orig.get("attribute"))
        key = name.lower()
        ai_row = by_name.get(key, {})
        cand = candidate_by_name.get(key, {})
        candidate_value = clean_text(ai_row.get("candidate_value")) or clean_text(cand.get("value"))
        status = _deterministic_status(clean_text(orig.get("value")), candidate_value) or clean_text(ai_row.get("status")) or "Not Confirmed"
        if status not in {"Match", "Equivalent", "Conflict", "Not Confirmed", "Not Applicable"}:
            status = "Not Confirmed"
        try:
            confidence = float(ai_row.get("confidence", 0.7))
        except (TypeError, ValueError):
            confidence = 0.7
        comparisons.append(AttributeComparison(
            category=clean_text(orig.get("category")) or "Other",
            attribute=name,
            original_value=clean_text(orig.get("value")),
            candidate_value=candidate_value,
            original_unit=clean_text(orig.get("unit")),
            candidate_unit=clean_text(cand.get("unit")),
            requirement_level=clean_text(orig.get("requirement_level")) or "Required",
            status=status,
            confidence=max(0.0, min(1.0, confidence)),
            explanation=clean_text(ai_row.get("explanation")),
            original_page=clean_text(orig.get("source_page")),
            candidate_page=clean_text(ai_row.get("candidate_page")) or clean_text(cand.get("source_page")),
        ))

    required = [c for c in comparisons if c.requirement_level.lower() == "required"]
    hard_conflicts = sum(c.status == "Conflict" for c in required)
    unconfirmed = sum(c.status == "Not Confirmed" for c in required)
    confirmed = sum(c.status in {"Match", "Equivalent"} for c in required)
    coverage = 100.0 * confirmed / max(1, len(required))
    weighted = {"Match": 1.0, "Equivalent": 0.92, "Not Applicable": 1.0, "Not Confirmed": 0.25, "Conflict": 0.0}
    score = 100.0 * sum(weighted[c.status] for c in required) / max(1, len(required))
    if hard_conflicts:
        status = "Not Compatible"
        score = min(score, 49.0)
    elif unconfirmed:
        status = "Needs Verification"
        score = min(score, 89.0)
    elif required and all(c.status == "Match" for c in required):
        status = "Exact Specification Match"
    else:
        status = "Technical Equivalent"

    return CandidateComparison(
        candidate_name=clean_text(candidate.get("document_title")) or clean_text(candidate.get("filename")),
        manufacturer=clean_text(candidate.get("manufacturer")),
        model=clean_text(candidate.get("model")),
        status=status,
        compatibility_score=round(score, 1),
        evidence_coverage=round(coverage, 1),
        hard_conflicts=hard_conflicts,
        unconfirmed_required=unconfirmed,
        comparisons=comparisons,
        summary=clean_text(ai.get("summary")),
    )


def comparison_rows(result: CandidateComparison) -> list[dict[str, Any]]:
    return [asdict(row) for row in result.comparisons]


def create_spec_comparison_workbook(original: dict[str, Any], results: Iterable[CandidateComparison]) -> tuple[str, bytes]:
    results = list(results)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Summary"
    dark = "17324D"; blue = "2E75B6"; light = "DCE6F1"; green = "E2F0D9"; red = "FCE4D6"; yellow = "FFF2CC"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "SPECIFICATION COMPARISON REPORT"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.append(["Original document", original.get("document_title") or original.get("filename"), "Manufacturer", original.get("manufacturer"), "Model", original.get("model"), "Attributes", len(original.get("attributes", []))])
    ws.append([])
    headers = ["Candidate", "Manufacturer", "Model", "Status", "Compatibility", "Evidence coverage", "Hard conflicts", "Unconfirmed required"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=blue)
    for result in sorted(results, key=lambda r: (r.hard_conflicts, -r.compatibility_score)):
        ws.append([result.candidate_name, result.manufacturer, result.model, result.status, result.compatibility_score / 100, result.evidence_coverage / 100, result.hard_conflicts, result.unconfirmed_required])
    for row in range(5, ws.max_row + 1):
        ws.cell(row, 5).number_format = "0.0%"; ws.cell(row, 6).number_format = "0.0%"
        status = str(ws.cell(row, 4).value)
        fill = green if status in {"Exact Specification Match", "Technical Equivalent"} else red if status == "Not Compatible" else yellow
        ws.cell(row, 4).fill = PatternFill("solid", fgColor=fill)
    widths = [34, 20, 20, 26, 15, 17, 14, 20]
    for i, width in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A5"

    detail = wb.create_sheet("Spec Comparison")
    detail.sheet_view.showGridLines = False
    dheaders = ["Candidate", "Category", "Attribute", "Requirement", "Original value", "Candidate value", "Status", "Confidence", "Explanation", "Original page", "Candidate page"]
    detail.append(dheaders)
    for cell in detail[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=dark)
    for result in results:
        for c in result.comparisons:
            detail.append([result.candidate_name, c.category, c.attribute, c.requirement_level, c.original_value, c.candidate_value, c.status, c.confidence, c.explanation, c.original_page, c.candidate_page])
    for row in range(2, detail.max_row + 1):
        detail.cell(row, 8).number_format = "0%"
        status = str(detail.cell(row, 7).value)
        detail.cell(row, 7).fill = PatternFill("solid", fgColor=green if status in {"Match", "Equivalent"} else red if status == "Conflict" else yellow)
        for col in range(1, 12): detail.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
    for i, width in enumerate([30, 16, 26, 14, 32, 32, 18, 12, 48, 14, 14], 1): detail.column_dimensions[get_column_letter(i)].width = width
    detail.freeze_panes = "A2"; detail.auto_filter.ref = detail.dimensions

    evidence = wb.create_sheet("Original Evidence")
    evidence.append(["Category", "Attribute", "Value", "Unit", "Requirement", "Page", "Evidence"])
    for cell in evidence[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=dark)
    for attr in original.get("attributes", []):
        evidence.append([attr.get("category"), attr.get("attribute"), attr.get("value"), attr.get("unit"), attr.get("requirement_level"), attr.get("source_page"), attr.get("evidence")])
    for i, width in enumerate([16, 28, 35, 12, 14, 12, 60], 1): evidence.column_dimensions[get_column_letter(i)].width = width
    for row in evidence.iter_rows():
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)

    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", clean_text(original.get("model")) or clean_text(original.get("document_title")) or "Spec_Comparison").strip("_")[:70]
    out = BytesIO(); wb.save(out)
    return f"{safe}_Spec_Comparison.xlsx", out.getvalue()
