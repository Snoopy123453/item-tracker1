from io import BytesIO

from openpyxl import Workbook, load_workbook

from product_finder.purchase_tracker import extract_purchase_candidates, create_purchase_tracker_bytes


def _source_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Results"
    ws.append(["title", "seller", "extracted_price", "product_link", "thumbnail", "query"])
    ws.append(["Test Sink", "Supply Store", 125.50, "https://example.com/sink", "https://example.com/image.png", "TEST-100"])
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_extract_purchase_candidates():
    rows = extract_purchase_candidates(_source_workbook())
    assert len(rows) == 1
    assert rows[0]["product"] == "Test Sink"
    assert rows[0]["unit_price"] == 125.50
    assert rows[0]["product_link"] == "https://example.com/sink"


def test_create_purchase_tracker():
    rows = extract_purchase_candidates(_source_workbook())
    rows[0]["quantity"] = 2
    filename, payload = create_purchase_tracker_bytes(rows, tracker_name="Test Job Purchases")
    assert filename == "Test_Job_Purchases.xlsx"
    wb = load_workbook(BytesIO(payload), data_only=False)
    assert wb.sheetnames == ["Dashboard", "Purchase List", "Instructions"]
    ws = wb["Purchase List"]
    assert ws["K2"].value == "=G2*H2+I2+J2"
    assert ws["E2"].hyperlink.target == "https://example.com/sink"
