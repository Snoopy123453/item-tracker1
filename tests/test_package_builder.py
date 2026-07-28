from io import BytesIO
from openpyxl import load_workbook

from product_finder.package_builder import build_known_package, create_package_workbook, package_to_rfq_items


def test_s1_package_expands_to_six_components():
    package = build_known_package("S-1", "JUST USXN1842A-J complete with Chicago 350-GN8AE35ABCP")
    assert package is not None
    assert len(package.components) == 6
    assert package.components[0].model == "USXN1842A-J"
    assert package.components[-1].model == "PW2150GJ"


def test_package_rfq_quantities_scale():
    package = build_known_package("S-1", "USXN1842A-J 350-GN8AE35ABCP", quantity=2)
    rows = package_to_rfq_items(package)
    assert len(rows) == 6
    assert all(row["quantity"] == 2 for row in rows)


def test_package_workbook_has_lead_time_and_quote_fields():
    package = build_known_package("S-1", "USXN1842A-J 350-GN8AE35ABCP")
    wb = load_workbook(BytesIO(create_package_workbook(package)), data_only=False)
    ws = wb["Package Summary"]
    headers = [cell.value for cell in ws[8]]
    assert "Lead Time" in headers
    assert "Quote #" in headers
    assert "Package Total" == ws.cell(16, 8).value
