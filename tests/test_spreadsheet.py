from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from product_finder.models import InputRecord, ProductResult, StoreResult
from product_finder.spreadsheet import create_product_workbook_bytes


def test_workbook_is_valid_and_escapes_formula_text() -> None:
    filename, data = create_product_workbook_bytes(
        input_records=[
            InputRecord(
                input_type="text",
                label="=HYPERLINK(\"https://bad.example\",\"click\")",
                generated_queries=["test product"],
            )
        ],
        product_results=[
            ProductResult(
                query="test product",
                input_source="typed",
                rank=1,
                title="Sample Product",
                seller="Retailer",
                price="$10.00",
                extracted_price=10.0,
                product_link="https://example.com/item",
            )
        ],
        store_results=[
            StoreResult(
                query="test product",
                rank=1,
                title="Sample Store",
                address="123 Main St",
                maps_link="https://maps.example.com/store",
            )
        ],
        location="90001",
    )

    assert filename.endswith(".xlsx")
    workbook = load_workbook(BytesIO(data), data_only=False)
    assert workbook.sheetnames == ["Summary", "Inputs", "Product Results", "Nearby Stores"]
    assert workbook["Inputs"]["B2"].value.startswith("'")
    assert workbook["Product Results"]["M2"].hyperlink.target == "https://example.com/item"
    assert workbook["Nearby Stores"]["L2"].hyperlink.target == "https://maps.example.com/store"
