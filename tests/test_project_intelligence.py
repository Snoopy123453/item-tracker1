from io import BytesIO
import json
from openpyxl import load_workbook

from product_finder.project_intelligence import consolidate_items, create_project_backup, load_project_backup, create_project_workbook


def test_consolidates_same_model_and_sums_quantity():
    rows = [
        {"item_tag":"FD-1","manufacturer":"JOSAM","model":"30000-5A-Z","description":"Floor drain","quantity":2,"source_file":"P1.pdf"},
        {"item_tag":"FD-2","manufacturer":"Josam","model":"30000-5A-Z","description":"Floor drain","quantity":3,"source_file":"P2.pdf"},
    ]
    out = consolidate_items(rows)
    assert len(out) == 1
    assert out[0]["quantity"] == 5
    assert "FD-1" in out[0]["item_tag"] and "FD-2" in out[0]["item_tag"]


def test_backup_round_trip():
    project = {"project_name":"Test Job","equipment":[],"preferences":{}}
    name, data = create_project_backup(project)
    assert name.endswith(".json")
    assert load_project_backup(data)["project_name"] == "Test Job"


def test_project_workbook_has_register():
    project = {"project_name":"Test Job","equipment":[{"item_tag":"S-1","manufacturer":"JUST","model":"USXN1842A-J","description":"Sink","quantity":1,"status":"Approved"}],"preferences":{"require_exact_model":True}}
    name, data = create_project_workbook(project)
    wb = load_workbook(BytesIO(data), data_only=False)
    assert name == "Test_Job_Procurement.xlsx"
    assert "Equipment Register" in wb.sheetnames
    assert wb["Equipment Register"]["C2"].value == "JUST"
