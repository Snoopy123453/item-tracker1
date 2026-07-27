from product_finder.procurement_controls import (
    compare_requirements, package_completeness, landed_cost,
    group_duplicate_offers, normalize_vendor, classify_document,
    data_health_checks, build_review_queue, create_procurement_control_workbook,
)
from openpyxl import load_workbook
from io import BytesIO


def test_hard_requirement_rejects_conflict():
    req=[{"attribute":"model","required_value":"30000-5A-Z","importance":"Required","weight":1.0}]
    comparisons,reject,score=compare_requirements(req,{"model":"30003-Z-5A"})
    assert reject is True
    assert comparisons[0].status == "Conflict"
    assert score == 0


def test_package_and_landed_cost():
    result=package_completeness(["sink","faucet","strainer"],"Sink with faucet")
    assert result["percent"] == 66.7
    calc=landed_cost(100,2,shipping=10,tax_rate=.10,discount=5,accessory_cost=20)
    assert calc["delivered_total"] == 246.5


def test_duplicate_grouping_and_vendor_normalization():
    rows=[{"title":"A","manufacturer":"JOSAM","model":"30000","seller":"The Home Depot"},{"title":"A second","manufacturer":"JOSAM","model":"30000","seller":"homedepot.com"}]
    groups=group_duplicate_offers(rows)
    assert len(groups)==1 and groups[0]["offer_count"]==2
    assert normalize_vendor("The Home Depot") == "Home Depot"


def test_document_classification_and_review():
    assert classify_document("Installation Instructions", "x.pdf") == "Installation Manual"
    rows=[{"title":"Thing","model":"X1","seller":"","match_score":70,"quantity":0,"product_link":""}]
    assert build_review_queue(rows,85)
    issues=data_health_checks(rows)
    assert len(issues)>=3


def test_control_workbook():
    name,data=create_procurement_control_workbook("Demo",[{"title":"Valve","model":"V1","seller":"Grainger","quantity":2,"unit_price":10,"status":"Approved","approved":True}],[],[],[])
    assert name.endswith('.xlsx')
    wb=load_workbook(BytesIO(data),data_only=False)
    assert {"Control Dashboard","Products","Review Queue","Data Health","PO Draft"}.issubset(wb.sheetnames)
