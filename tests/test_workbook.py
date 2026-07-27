from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from product_finder.models import InputRecord, ProductResult, StoreResult
from product_finder.spreadsheet import create_product_workbook_bytes


def test_workbook_structure_and_formula_safety() -> None:
    _, workbook_bytes = create_product_workbook_bytes(
        input_records=[
            InputRecord(
                input_type="text",
                label="=2+2",
                extracted_product_name="test product",
                generated_queries=["test product"],
            )
        ],
        product_results=[
            ProductResult(
                query="test product",
                input_source="=2+2",
                rank=1,
                title="@SUM(A1:A2)",
                seller="Example Store",
                price="$19.99",
                extracted_price=19.99,
                product_link="https://example.com/product",
            )
        ],
        store_results=[
            StoreResult(
                query="test product",
                rank=1,
                title="+Example Retailer",
                address="123 Main St",
                maps_link="https://example.com/map",
            )
        ],
        location="90210",
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert workbook.sheetnames == ["Summary", "Inputs", "Product Results", "Nearby Stores"]
    assert workbook["Summary"]["B6"].value == 1
    assert workbook["Summary"]["B7"].value == 1
    assert workbook["Summary"]["B8"].value == 1

    assert workbook["Inputs"]["B2"].value == "'=2+2"
    assert workbook["Product Results"]["D2"].value == "'@SUM(A1:A2)"
    assert workbook["Nearby Stores"]["C2"].value == "'+Example Retailer"
    assert workbook["Product Results"]["M2"].hyperlink.target == "https://example.com/product"

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                assert cell.data_type != "f"
