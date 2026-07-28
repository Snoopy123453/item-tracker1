from io import BytesIO
from openpyxl import Workbook, load_workbook
from product_finder.rfq_builder import extract_rfq_items, build_rfq_email, create_rfq_workbook

def test_extract_and_build_rfq():
    wb=Workbook(); ws=wb.active; ws.title='Equipment Register'
    ws.append(['Item','Manufacturer','Model','Description','Quantity','Product Link'])
    ws.append(['FD-1','JOSAM','30002-5A-Z-50','Floor drain',2,'https://example.com'])
    out=BytesIO(); wb.save(out)
    items=extract_rfq_items(out.getvalue())
    assert len(items)==1 and items[0].model=='30002-5A-Z-50'
    rows=[items[0].to_row()]
    email=build_rfq_email('Test','90802','2026-08-01','Daniel','d@example.com',rows,'No substitutions',False)
    assert 'lead time' in email.lower() and '30002-5A-Z-50' in email
    data=create_rfq_workbook({'project_name':'Test','email_draft':email},rows)
    wb2=load_workbook(BytesIO(data),data_only=False)
    assert 'RFQ' in wb2.sheetnames and 'Email Draft' in wb2.sheetnames
